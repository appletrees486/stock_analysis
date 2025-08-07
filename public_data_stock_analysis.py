#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
공공데이터 API를 사용한 국내 주식 일봉 시세 조회 스크립트
"""

# matplotlib 백엔드를 Agg로 설정 (tkinter 에러 방지)
import matplotlib
matplotlib.use('Agg')

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import requests
import json
import platform
import os
import time
# openpyxl import 추가
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 운영체제별 한글 폰트 설정
system = platform.system()
if system == 'Windows':
    # Windows 환경
    font_list = ['Malgun Gothic', '맑은 고딕', 'NanumGothic', '나눔고딕']
elif system == 'Darwin':  # macOS
    font_list = ['AppleGothic', 'NanumGothic', '나눔고딕']
else:  # Linux
    font_list = ['NanumGothic', '나눔고딕', 'DejaVu Sans']

# 사용 가능한 폰트 찾기
available_font = None
for font in font_list:
    try:
        fm.findfont(font)
        available_font = font
        break
    except:
        continue

if available_font:
    plt.rcParams['font.family'] = available_font
    print(f"✅ 사용 폰트: {available_font}")
else:
    # 기본 폰트 사용
    plt.rcParams['font.family'] = 'DejaVu Sans'
    print("⚠️ 한글 폰트를 찾을 수 없어 기본 폰트를 사용합니다.")

plt.rcParams['axes.unicode_minus'] = False

# 폰트 캐시 재설정
try:
    fm._rebuild()
except AttributeError:
    # 최신 matplotlib 버전에서는 _rebuild가 제거됨
    fm.findfont('DejaVu Sans', rebuild_if_missing=True)

# 공공데이터 API 설정
PUBLIC_DATA_API_KEY = "1A4ObzoWfk8XusMwdYIzhsp4F/vPMPCPtuRvV+VJnKsI1YrCzEmwOByBy/Vw1MMbEkNUxVQrZr63qd21UM0sYw=="
PUBLIC_DATA_BASE_URL = "http://apis.data.go.kr/1160100/service/GetStockSecuritiesInfoService"

# 종목번호와 종목명 매핑 딕셔너리
STOCK_NAME_MAPPING = {
    '005930': '삼성전자',
    '000660': 'SK하이닉스',
    '035420': 'NAVER',
    '051910': 'LG화학',
    '006400': '삼성SDI',
    '035720': '카카오',
    '207940': '삼성바이오로직스',
    '068270': '셀트리온',
    '323410': '카카오뱅크',
    '005380': '현대차',
    '051900': 'LG생활건강',
    '373220': 'LG에너지솔루션',
    '000270': '기아',
    '259960': '크래프톤',
    '034730': 'SK',
    '017670': 'SK텔레콤',
    '011070': 'LG이노텍',
    '028260': '삼성물산',
    '096770': 'SK이노베이션',
    '018260': '삼성에스디에스',
    '079160': 'CJ CGV',
    '035760': 'CJ ENM',
    '311690': 'CJ 바이오사이언스',
    '000120': 'CJ대한통운',
    '097950': 'CJ제일제당',
    '051500': 'CJ프레시웨이',
    '011150': 'CJ대한통운',
    '039490': '키움증권',
    '039830': '오로라월드',
    '900270': '항성홀딩그룹',
    '380550': '뉴로펫',
    '328130': '루닛',
    '338220': 'Vuno',
    '145720': '덴티움',
    '014280': '금강킨드',
    '019210': 'YG-1',
    '023410': '유진기업',
    '012030': 'DB',
    '000990': 'DB하이텍',
    '005830': 'DB손해보험',
    '016610': 'DB금융투자',
    '058820': 'CMG제약',
    '056730': 'CNT85',
    '000480': 'CR홀딩스',
    '065770': 'CS',
    '000590': 'CS홀딩스',
    '083660': 'CSA코스믹',
    '477760': 'DB금융스팩12호',
    '060900': 'DGP',
}

def get_stock_name_from_code(stock_code):
    """종목번호로 종목명을 가져오는 함수"""
    return STOCK_NAME_MAPPING.get(stock_code, stock_code)

def get_current_date():
    """현재 날짜를 가져오는 함수 (시스템 날짜 문제 해결)"""
    current_date = datetime.now()
    
    # 시스템 날짜가 미래로 설정되어 있는지 확인
    if current_date.year > 2024:
        print(f"⚠️ 시스템 날짜가 미래로 설정되어 있습니다: {current_date.strftime('%Y-%m-%d')}")
        print(f"🔄 2024년 12월 27일로 수동 설정합니다.")
        return datetime(2024, 12, 27)
    else:
        return current_date

def find_latest_trading_date(stock_code, max_attempts=30):
    """최근 거래일을 찾는 함수"""
    print(f"🔍 최근 거래일을 찾는 중...")
    
    # 현재 날짜부터 시작
    current_date = get_current_date()
    
    for i in range(max_attempts):
        test_date = current_date - timedelta(days=i)
        test_date_str = test_date.strftime("%Y%m%d")
        
        print(f"   📅 {test_date.strftime('%Y-%m-%d')} 확인 중... ({i+1}/{max_attempts})")
        
        # API 호출
        params = {
            'serviceKey': PUBLIC_DATA_API_KEY,
            'pageNo': '1',
            'numOfRows': '1',  # 1개만 확인
            'resultType': 'json',
            'basDt': test_date_str,
            'srtnCd': stock_code
        }
        
        try:
            response = requests.get(
                f"{PUBLIC_DATA_BASE_URL}/getStockPriceInfo", 
                params=params,
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                if 'response' in data and 'body' in data['response']:
                    total_count = data['response']['body'].get('totalCount', 0)
                    if total_count > 0:
                        print(f"✅ 최근 거래일 발견: {test_date.strftime('%Y-%m-%d')}")
                        return test_date
            
            # 잠시 대기 (API 호출 제한 방지)
            time.sleep(0.5)
            
        except Exception as e:
            print(f"   ⚠️ {test_date.strftime('%Y-%m-%d')} 확인 중 오류: {str(e)}")
            continue
    
    print(f"❌ 최근 거래일을 찾을 수 없습니다. 기본값 사용: {current_date.strftime('%Y-%m-%d')}")
    return current_date

def get_stock_data(stock_code):
    """공공데이터 API를 사용한 국내 주식 일봉 데이터 조회 (240일)"""
    print(f"🔍 {stock_code} 240일 일봉 시세 조회 중... (공공데이터 API)")
    print("   📅 일봉 데이터는 거래일 기준으로 제공되며, 주말/공휴일은 포함되지 않습니다.")
    
    try:
        # 1단계: 최근 거래일 찾기
        end_date = find_latest_trading_date(stock_code)
        
        # 2단계: 최근 거래일 기준으로 한 번에 데이터 조회
        print(f"   📅 조회 기준일: {end_date.strftime('%Y-%m-%d')}")
        print(f"   🔄 공공데이터 API에서 데이터 조회 중...")
        
        # API 호출 (한 번에 최대 1000개 데이터)
        params = {
            'serviceKey': PUBLIC_DATA_API_KEY,
            'pageNo': '1',
            'numOfRows': '1000',  # 최대 1000개 행
            'resultType': 'json',
            'basDt': end_date.strftime("%Y%m%d"),  # 기준일자
            'srtnCd': stock_code    # 종목코드
        }
        
        print(f"   🔍 API 요청 파라미터: basDt={end_date.strftime('%Y%m%d')}, srtnCd={stock_code}")
        
        try:
            response = requests.get(
                f"{PUBLIC_DATA_BASE_URL}/getStockPriceInfo", 
                params=params,
                timeout=30
            )
            
            print(f"   📊 응답 상태: {response.status_code}")
            
            if response.status_code == 200:
                try:
                    data = response.json()
                    
                    # 응답 데이터 구조 확인
                    print(f"   📋 API 응답 구조: {list(data.keys())}")
                    
                    # 성공적인 응답인지 확인
                    if 'response' in data and 'body' in data['response']:
                        items = data['response']['body'].get('items', {})
                        
                        if isinstance(items, dict) and 'item' in items:
                            items = items['item']
                        
                        if items and len(items) > 0:
                            print(f"✅ 공공데이터 API 일봉: {len(items)}개의 데이터를 조회했습니다.")
                            
                            # DataFrame으로 변환
                            df = pd.DataFrame(items)
                            
                            # 컬럼명 확인 및 매핑
                            print(f"   📋 사용 가능한 컬럼: {list(df.columns)}")
                            
                            # 필요한 컬럼 매핑
                            column_mapping = {
                                'basDt': 'Date',           # 기준일자
                                'srtnCd': 'StockCode',     # 종목코드
                                'itmsNm': 'StockName',     # 종목명
                                'clpr': 'Close',           # 종가
                                'vs': 'Change',            # 대비
                                'fltRt': 'ChangeRate',     # 등락률
                                'mkp': 'Open',             # 시가
                                'hipr': 'High',            # 고가
                                'lopr': 'Low',             # 저가
                                'trqu': 'Volume',          # 거래량
                                'trPrc': 'TradePrice',     # 거래대금
                                'lstgStCnt': 'ListedShares', # 상장주식수
                                'mrktTotAmt': 'MarketCap'   # 시가총액
                            }
                            
                            # 컬럼명 변경
                            df = df.rename(columns=column_mapping)
                            
                            # 종목명 매칭 - 매핑 딕셔너리 우선
                            stock_name = get_stock_name_from_code(stock_code)
                            df['StockName'] = stock_name
                            
                            # 날짜 형식 변환
                            if 'Date' in df.columns:
                                df['Date'] = pd.to_datetime(df['Date'], format='%Y%m%d')
                                df.set_index('Date', inplace=True)
                            
                            # 숫자 컬럼 변환
                            numeric_columns = ['Close', 'Change', 'ChangeRate', 'Open', 'High', 'Low', 'Volume', 'TradePrice', 'ListedShares', 'MarketCap']
                            for col in numeric_columns:
                                if col in df.columns:
                                    df[col] = pd.to_numeric(df[col], errors='coerce')
                            
                            # 데이터 정렬 (최신순)
                            df.sort_index(inplace=True)
                            
                            # 240일 데이터로 제한
                            if len(df) > 240:
                                df = df.tail(240)
                            
                            print(f"📅 총 {len(df)}일의 일봉 거래 데이터를 가져왔습니다.")
                            print(f"🏢 종목명: {stock_name}")
                            
                            # 디버깅: 데이터 기간 확인
                            print(f"🔍 데이터 기간 디버깅:")
                            print(f"   요청 기간: 240일")
                            print(f"   실제 시작일: {df.index[0].strftime('%Y-%m-%d')}")
                            print(f"   실제 종료일: {df.index[-1].strftime('%Y-%m-%d')}")
                            print(f"   실제 데이터 수: {len(df)}일")
                            
                            # 최근 5일 데이터 출력
                            print(f"📊 최근 일봉 데이터 상세:")
                            for i, (date, row) in enumerate(df.tail(5).iterrows()):
                                print(f"   {date.strftime('%Y-%m-%d')}: {row.get('Open', 0):,.0f} → {row.get('Close', 0):,.0f} (거래량: {row.get('Volume', 0):,.0f})")
                            
                            return df
                        else:
                            print(f"   ⚠️ 데이터가 없습니다")
                            print(f"   💡 해당 종목코드({stock_code})의 데이터가 없거나 조회 조건을 확인해주세요.")
                            print(f"   🔍 API 응답 내용: {data}")
                            return None
                    else:
                        print(f"   ⚠️ 응답 형식이 올바르지 않습니다")
                        print(f"   📋 응답 내용: {data}")
                        return None
                        
                except json.JSONDecodeError:
                    print(f"   ❌ JSON 파싱 실패")
                    print(f"   📋 응답 내용: {response.text[:200]}...")
                    return None
            else:
                print(f"   ❌ HTTP 오류: {response.status_code}")
                print(f"   📋 응답 내용: {response.text[:200]}...")
                return None
                
        except requests.exceptions.RequestException as e:
            print(f"   ❌ 요청 오류: {str(e)}")
            return None
        
    except Exception as e:
        print(f"   ❌ 공공데이터 API 연결 실패: {str(e)}")
        return None
    
    # 모든 소스에서 실패
    print("❌ 일봉 데이터 조회에 실패했습니다.")
    print("💡 가능한 원인:")
    print("   - 종목코드가 잘못되었습니다")
    print("   - 해당 종목이 상장폐지되었습니다")
    print("   - 네트워크 연결에 문제가 있습니다")
    return None

def calculate_technical_indicators(df):
    """기술적 지표 계산"""
    # 이동평균선
    df['MA5'] = df['Close'].rolling(window=5).mean()
    df['MA20'] = df['Close'].rolling(window=20).mean()
    df['MA60'] = df['Close'].rolling(window=60).mean()
    df['MA120'] = df['Close'].rolling(window=120).mean()
    
    # MACD 계산 (표준 공식)
    # MACD = 12일 EMA - 26일 EMA
    # Signal = MACD의 9일 EMA
    # Histogram = MACD - Signal
    ema12 = df['Close'].ewm(span=12, adjust=False).mean()
    ema26 = df['Close'].ewm(span=26, adjust=False).mean()
    df['MACD'] = ema12 - ema26
    df['MACD_Signal'] = df['MACD'].ewm(span=9, adjust=False).mean()
    df['MACD_Histogram'] = df['MACD'] - df['MACD_Signal']
    
    # RSI 계산 (표준 공식)
    # RSI = 100 - (100 / (1 + RS))
    # RS = 평균 상승폭 / 평균 하락폭
    delta = df['Close'].diff()
    
    # 상승폭과 하락폭 분리
    gain = delta.copy()
    loss = delta.copy()
    gain[gain < 0] = 0
    loss[loss > 0] = 0
    loss = abs(loss)
    
    # 14일 평균 계산
    avg_gain = gain.rolling(window=14).mean()
    avg_loss = loss.rolling(window=14).mean()
    
    # RS와 RSI 계산
    rs = avg_gain / avg_loss
    df['RSI'] = 100 - (100 / (1 + rs))
    
    return df

def analyze_stock_data(hist, stock_code):
    """주식 일봉 데이터 분석"""
    if hist is None or hist.empty:
        return
    
    print("\n" + "="*60)
    print(f"📊 {stock_code} 주식 일봉 분석 결과 (공공데이터 API)")
    print("="*60)
    
    # 기본 통계
    print(f"📅 조회 기간: {hist.index[0].strftime('%Y-%m-%d')} ~ {hist.index[-1].strftime('%Y-%m-%d')}")
    print(f"📈 일봉 거래일 수: {len(hist)}일")
    
    # 가격 정보
    print(f"\n💰 가격 정보:")
    print(f"   시작가: {hist['Open'].iloc[0]:,.0f}원")
    print(f"   종가: {hist['Close'].iloc[-1]:,.0f}원")
    print(f"   최고가: {hist['High'].max():,.0f}원")
    print(f"   최저가: {hist['Low'].min():,.0f}원")
    
    # 변동 정보
    price_change = hist['Close'].iloc[-1] - hist['Open'].iloc[0]
    price_change_pct = (price_change / hist['Open'].iloc[0]) * 100
    
    print(f"\n📊 변동 정보:")
    print(f"   가격 변동: {price_change:+,.0f}원")
    print(f"   변동률: {price_change_pct:+.2f}%")
    
    # 일봉 거래량 정보
    print(f"\n📈 일봉 거래량 정보:")
    print(f"   평균 일봉 거래량: {hist['Volume'].mean():,.0f}주")
    print(f"   최대 일봉 거래량: {hist['Volume'].max():,.0f}주")
    print(f"   최소 일봉 거래량: {hist['Volume'].min():,.0f}주")
    
    # 기술적 지표 계산
    df_with_indicators = calculate_technical_indicators(hist.copy())
    
    # 기술적 지표 정보
    print(f"\n📊 기술적 지표 (최근값):")
    print(f"   5일 이동평균: {df_with_indicators['MA5'].iloc[-1]:,.0f}원")
    print(f"   20일 이동평균: {df_with_indicators['MA20'].iloc[-1]:,.0f}원")
    print(f"   60일 이동평균: {df_with_indicators['MA60'].iloc[-1]:,.0f}원")
    print(f"   120일 이동평균: {df_with_indicators['MA120'].iloc[-1]:,.0f}원")
    
    # RSI 정보
    rsi_value = df_with_indicators['RSI'].iloc[-1]
    print(f"   RSI: {rsi_value:.1f}")
    if rsi_value > 70:
        print("   RSI 신호: 과매수 구간")
    elif rsi_value < 30:
        print("   RSI 신호: 과매도 구간")
    else:
        print("   RSI 신호: 중립 구간")
    
    # MACD 정보
    macd_value = df_with_indicators['MACD'].iloc[-1]
    macd_signal = df_with_indicators['MACD_Signal'].iloc[-1]
    macd_histogram = df_with_indicators['MACD_Histogram'].iloc[-1]
    print(f"   MACD: {macd_value:.2f}")
    print(f"   MACD Signal: {macd_signal:.2f}")
    print(f"   MACD Histogram: {macd_histogram:.2f}")
    if macd_value > macd_signal:
        print("   MACD 신호: 상승 추세")
    else:
        print("   MACD 신호: 하락 추세")

def create_stock_chart(hist, stock_code):
    """주식 일봉 차트 생성 (캔들차트 + 보조지표) - 차트 데이터 반환 추가"""
    if hist is None or hist.empty:
        return None, None
    
    print(f"\n📈 일봉 캔들차트를 생성합니다... (공공데이터 API)")
    
    # 기술적 지표 계산
    df = calculate_technical_indicators(hist.copy())
    df.index.name = 'Date'
    
    # 종목명 가져오기
    stock_name = df['StockName'].iloc[0] if 'StockName' in df.columns else get_stock_name_from_code(stock_code)
    
    # 하나의 큰 차트에 모든 지표 포함
    fig, axes = plt.subplots(4, 1, figsize=(15, 16), height_ratios=[6, 2, 2, 2])
    fig.suptitle(f'{stock_name} ({stock_code}) Daily Stock Chart (240 Days) - Public Data API', fontsize=16, fontweight='bold')
    
    # 1. 캔들차트 (첫 번째 패널)
    ax1 = axes[0]
    
    # 캔들차트 그리기
    for date, row in df.iterrows():
        color = 'red' if row['Close'] >= row['Open'] else 'blue'
        ax1.plot([date, date], [row['Low'], row['High']], color=color, linewidth=1)
        ax1.plot([date, date], [row['Open'], row['Close']], color=color, linewidth=3)
    
    # 이동평균선 추가
    ax1.plot(df.index, df['MA5'], color='red', linewidth=1, alpha=0.7, label='MA5')
    ax1.plot(df.index, df['MA20'], color='green', linewidth=1, alpha=0.7, label='MA20')
    ax1.plot(df.index, df['MA60'], color='orange', linewidth=1, alpha=0.7, label='MA60')
    ax1.plot(df.index, df['MA120'], color='purple', linewidth=1, alpha=0.7, label='MA120')
    
    ax1.set_title('Price Chart with Moving Averages (Public Data API)')
    ax1.set_ylabel('Price (KRW)')
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # 2. 일봉 거래량 차트 (두 번째 패널)
    ax2 = axes[1]
    ax2.bar(df.index, df['Volume'], color='green', alpha=0.7)
    ax2.set_title('Daily Volume')
    ax2.set_ylabel('Volume')
    ax2.grid(True, alpha=0.3)
    
    # 3. MACD 차트 (세 번째 패널)
    ax3 = axes[2]
    ax3.plot(df.index, df['MACD'], color='blue', linewidth=1, label='MACD')
    ax3.plot(df.index, df['MACD_Signal'], color='red', linewidth=1, label='Signal')
    ax3.bar(df.index, df['MACD_Histogram'], color='gray', alpha=0.5, width=0.8, label='Histogram')
    ax3.axhline(y=0, color='black', linestyle='-', alpha=0.3)
    ax3.set_title('MACD')
    ax3.legend()
    ax3.grid(True, alpha=0.3)
    
    # 4. RSI 차트 (네 번째 패널)
    ax4 = axes[3]
    ax4.plot(df.index, df['RSI'], color='purple', linewidth=1, label='RSI')
    ax4.axhline(y=70, color='red', linestyle='--', alpha=0.5, label='Overbought')
    ax4.axhline(y=30, color='green', linestyle='--', alpha=0.5, label='Oversold')
    ax4.axhline(y=50, color='black', linestyle='-', alpha=0.3)
    ax4.set_ylim(0, 100)
    ax4.set_title('RSI')
    ax4.legend()
    ax4.grid(True, alpha=0.3)
    
    # 모든 패널의 x축 날짜 설정
    for ax in axes:
        # 날짜 인덱스에서 적절한 간격으로 날짜 선택
        date_indices = [df.index[0], df.index[len(df)//4], df.index[len(df)//2], 
                       df.index[3*len(df)//4], df.index[-1]]
        ax.set_xticks(date_indices)
        ax.set_xticklabels([date.strftime('%Y-%m-%d') for date in date_indices], 
                          rotation=45, ha='right')
    
    plt.tight_layout()
    
    # 차트를 이미지로 저장
    charts_dir = "public_data_charts"
    if not os.path.exists(charts_dir):
        os.makedirs(charts_dir)
        print(f"📁 {charts_dir} 폴더를 생성했습니다.")
    
    # 파일명 생성 (현재 날짜 사용)
    current_date = get_current_date().strftime("%Y%m%d")
    base_filename = f"public_data_{stock_name}_{stock_code}_{current_date}.png"
    
    # 파일명에서 특수문자 제거 및 공백을 언더스코어로 변경
    base_filename = base_filename.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_")
    
    # 파일 중복 확인 및 버전 추가
    version = 1
    filename = base_filename
    filepath = os.path.join(charts_dir, filename)
    
    while os.path.exists(filepath):
        # 파일명에서 확장자 분리
        name_without_ext = base_filename.rsplit('.', 1)[0]
        ext = base_filename.rsplit('.', 1)[1]
        filename = f"{name_without_ext}_v{version}.{ext}"
        filepath = os.path.join(charts_dir, filename)
        version += 1
    
    plt.savefig(filepath, dpi=150, bbox_inches='tight')
    print(f"💾 차트가 저장되었습니다: {filepath}")
    
    # 차트 뷰어를 띄우지 않고 차트 닫기
    plt.close()
    
    # 차트 데이터 반환 (보조지표 포함)
    return filepath, df

def save_chart_data_to_excel(chart_data, stock_code, stock_name):
    """차트 데이터를 엑셀로 저장 (보조지표 포함) - 공공데이터 API용"""
    if chart_data is None or chart_data.empty:
        print("❌ 저장할 차트 데이터가 없습니다.")
        return None
    
    try:
        print(f"\n📊 차트 데이터를 엑셀로 저장합니다... (공공데이터 API)")
        
        # 엑셀 파일 저장 디렉토리 생성
        excel_dir = "public_data_chart_excel"
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)
            print(f"📁 {excel_dir} 폴더를 생성했습니다.")
        
        # 파일명 생성 (현재 날짜 사용)
        current_date = get_current_date().strftime("%Y%m%d")
        base_filename = f"public_data_chart_data_{stock_name}_{stock_code}_{current_date}.xlsx"
        base_filename = base_filename.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_")
        
        # 파일 중복 확인 및 버전 추가
        version = 1
        filename = base_filename
        filepath = os.path.join(excel_dir, filename)
        
        while os.path.exists(filepath):
            name_without_ext = base_filename.rsplit('.', 1)[0]
            ext = base_filename.rsplit('.', 1)[1]
            filename = f"{name_without_ext}_v{version}.{ext}"
            filepath = os.path.join(excel_dir, filename)
            version += 1
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        
        # 기본 시트 제거
        wb.remove(wb.active)
        
        # 1. 기본 OHLCV 데이터 시트
        ws_basic = wb.create_sheet("기본데이터")
        
        # 기본 데이터 (OHLCV)
        basic_columns = ['Open', 'High', 'Low', 'Close', 'Volume']
        available_columns = [col for col in basic_columns if col in chart_data.columns]
        basic_data = chart_data[available_columns].copy()
        basic_data.index.name = 'Date'
        
        # 헤더 추가
        basic_data.insert(0, 'Date', basic_data.index)
        
        # 데이터프레임을 워크시트에 추가
        for r in dataframe_to_rows(basic_data, index=False, header=True):
            ws_basic.append(r)
        
        # 헤더 스타일링
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws_basic[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 컬럼 너비 조정
        ws_basic.column_dimensions['A'].width = 15  # Date
        for i, col in enumerate(available_columns, 1):
            ws_basic.column_dimensions[chr(65 + i)].width = 12
        
        # 2. 이동평균선 데이터 시트
        ws_ma = wb.create_sheet("이동평균선")
        
        ma_columns = ['Close', 'MA5', 'MA20', 'MA60', 'MA120']
        available_ma_columns = [col for col in ma_columns if col in chart_data.columns]
        ma_data = chart_data[available_ma_columns].copy()
        ma_data.index.name = 'Date'
        ma_data.insert(0, 'Date', ma_data.index)
        
        for r in dataframe_to_rows(ma_data, index=False, header=True):
            ws_ma.append(r)
        
        # 헤더 스타일링
        for cell in ws_ma[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 컬럼 너비 조정
        ws_ma.column_dimensions['A'].width = 15  # Date
        for i, col in enumerate(available_ma_columns, 1):
            ws_ma.column_dimensions[chr(65 + i)].width = 12
        
        # 3. MACD 데이터 시트
        ws_macd = wb.create_sheet("MACD")
        
        macd_columns = ['Close', 'MACD', 'MACD_Signal', 'MACD_Histogram']
        available_macd_columns = [col for col in macd_columns if col in chart_data.columns]
        macd_data = chart_data[available_macd_columns].copy()
        macd_data.index.name = 'Date'
        macd_data.insert(0, 'Date', macd_data.index)
        
        for r in dataframe_to_rows(macd_data, index=False, header=True):
            ws_macd.append(r)
        
        # 헤더 스타일링
        for cell in ws_macd[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 컬럼 너비 조정
        ws_macd.column_dimensions['A'].width = 15  # Date
        for i, col in enumerate(available_macd_columns, 1):
            ws_macd.column_dimensions[chr(65 + i)].width = 12
        
        # 4. RSI 데이터 시트
        ws_rsi = wb.create_sheet("RSI")
        
        rsi_columns = ['Close', 'RSI']
        available_rsi_columns = [col for col in rsi_columns if col in chart_data.columns]
        rsi_data = chart_data[available_rsi_columns].copy()
        rsi_data.index.name = 'Date'
        rsi_data.insert(0, 'Date', rsi_data.index)
        
        for r in dataframe_to_rows(rsi_data, index=False, header=True):
            ws_rsi.append(r)
        
        # 헤더 스타일링
        for cell in ws_rsi[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 컬럼 너비 조정
        ws_rsi.column_dimensions['A'].width = 15  # Date
        for i, col in enumerate(available_rsi_columns, 1):
            ws_rsi.column_dimensions[chr(65 + i)].width = 12
        
        # 5. 종합 데이터 시트 (모든 지표 포함)
        ws_summary = wb.create_sheet("종합데이터")
        
        # 모든 컬럼 선택
        summary_data = chart_data.copy()
        summary_data.index.name = 'Date'
        summary_data.insert(0, 'Date', summary_data.index)
        
        for r in dataframe_to_rows(summary_data, index=False, header=True):
            ws_summary.append(r)
        
        # 헤더 스타일링
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 컬럼 너비 조정
        for col in ws_summary.columns:
            ws_summary.column_dimensions[col[0].column_letter].width = 12
        
        # 6. 요약 정보 시트
        ws_info = wb.create_sheet("요약정보")
        
        # 기본 정보
        info_data = [
            ["종목명", stock_name],
            ["종목번호", stock_code],
            ["데이터소스", "공공데이터 API"],
            ["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["데이터 기간", f"{chart_data.index[0].strftime('%Y-%m-%d')} ~ {chart_data.index[-1].strftime('%Y-%m-%d')}"],
            ["총 데이터 수", len(chart_data)],
            ["", ""],
            ["최근 데이터 요약", ""],
        ]
        
        # 최근 데이터 정보 추가
        if 'Close' in chart_data.columns:
            info_data.append(["최근 종가", f"{chart_data['Close'].iloc[-1]:,.0f}원"])
        if 'RSI' in chart_data.columns:
            info_data.append(["최근 RSI", f"{chart_data['RSI'].iloc[-1]:.1f}"])
        if 'MACD' in chart_data.columns:
            info_data.append(["최근 MACD", f"{chart_data['MACD'].iloc[-1]:.2f}"])
        if 'MA5' in chart_data.columns:
            info_data.append(["5일 이동평균", f"{chart_data['MA5'].iloc[-1]:,.0f}원"])
        if 'MA20' in chart_data.columns:
            info_data.append(["20일 이동평균", f"{chart_data['MA20'].iloc[-1]:,.0f}원"])
        if 'MA60' in chart_data.columns:
            info_data.append(["60일 이동평균", f"{chart_data['MA60'].iloc[-1]:,.0f}원"])
        if 'MA120' in chart_data.columns:
            info_data.append(["120일 이동평균", f"{chart_data['MA120'].iloc[-1]:,.0f}원"])
        
        for row in info_data:
            ws_info.append(row)
        
        # 헤더 스타일링
        for row in ws_info.iter_rows(min_row=1, max_row=len(info_data)):
            for cell in row:
                if cell.value and cell.value in ["종목명", "종목번호", "데이터소스", "생성일시", "데이터 기간", "총 데이터 수", "최근 데이터 요약"]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 컬럼 너비 조정
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 30
        
        # 파일 저장
        wb.save(filepath)
        print(f"💾 엑셀 파일이 저장되었습니다: {filepath}")
        
        # 시트 정보 출력
        print(f"📊 생성된 시트:")
        print(f"   - 기본데이터: OHLCV 기본 정보")
        print(f"   - 이동평균선: 이동평균선 데이터")
        print(f"   - MACD: MACD 관련 지표")
        print(f"   - RSI: RSI 지표")
        print(f"   - 종합데이터: 모든 지표 통합")
        print(f"   - 요약정보: 종목 및 데이터 요약")
        
        return filepath
        
    except Exception as e:
        print(f"❌ 엑셀 파일 저장 중 오류: {e}")
        return None

def main():
    """메인 함수"""
    print("🚀 공공데이터 API를 사용한 국내 주식 일봉 시세 조회 프로그램 (240일)")
    print("="*60)
    
    # 종목코드 입력
    while True:
        stock_code = input("📈 종목코드를 입력하세요 (예: 005930): ").strip()
        if stock_code.isdigit() and len(stock_code) == 6:
            break
        else:
            print("❌ 올바른 종목코드를 입력해주세요 (6자리 숫자)")
    
    # 공공데이터 API 데이터 조회
    hist = get_stock_data(stock_code)
    
    if hist is not None:
        # 공공데이터 API 데이터 분석
        analyze_stock_data(hist, stock_code)
        
        # 공공데이터 API 차트 생성 (차트 데이터 반환)
        chart_path, chart_data = create_stock_chart(hist, stock_code)
        
        if chart_path and chart_data is not None:
            # 종목명 가져오기
            stock_name = chart_data['StockName'].iloc[0] if 'StockName' in chart_data.columns else get_stock_name_from_code(stock_code)
            
            # 차트 데이터를 엑셀로 저장
            excel_path = save_chart_data_to_excel(chart_data, stock_code, stock_name)
            
            if excel_path:
                print(f"\n✅ 공공데이터 API 분석이 완료되었습니다!")
                print(f"📈 차트 이미지: {chart_path}")
                print(f"📊 엑셀 데이터: {excel_path}")
                print(f"\n💡 이제 AI 분석에 차트 이미지와 엑셀 데이터를 함께 전달할 수 있습니다!")
            else:
                print(f"\n✅ 공공데이터 API 분석이 완료되었습니다!")
                print(f"📈 차트 이미지: {chart_path}")
                print(f"❌ 엑셀 데이터 저장에 실패했습니다.")
        else:
            print(f"\n❌ 차트 생성에 실패했습니다.")
    else:
        print("\n❌ 공공데이터 API 데이터 조회에 실패했습니다.")

if __name__ == "__main__":
    main() 