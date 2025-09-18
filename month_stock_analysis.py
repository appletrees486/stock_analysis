#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
국내 주식 월봉 시세 조회 스크립트 (DB 기반 + 공통 컨센선스 적용)
"""

# matplotlib 백엔드를 Agg로 설정 (tkinter 에러 방지)
import matplotlib
matplotlib.use('Agg')

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import mplfinance as mpf
import platform
import os
# openpyxl import 추가
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

# 데이터베이스 연결을 위한 import 추가
from database_config import DatabaseManager
# 향상된 데이터 검증 시스템 import 추가
from enhanced_data_validator import EnhancedDataValidator
# 한국 공휴일 관리자 import 추가
from korean_holiday_manager import KoreanHolidayManager

# 운영체제별 한글 폰트 설정 (한글 깨짐 방지 강화)
system = platform.system()
if system == 'Windows':
    # Windows 환경 - 한글 폰트 우선순위
    font_list = ['Malgun Gothic', '맑은 고딕', 'NanumGothic', '나눔고딕', 'Noto Sans CJK KR', 'Noto Sans KR']
elif system == 'Darwin':  # macOS
    font_list = ['AppleGothic', 'NanumGothic', '나눔고딕', 'Noto Sans CJK KR', 'Noto Sans KR']
else:  # Linux
    font_list = ['NanumGothic', '나눔고딕', 'Noto Sans CJK KR', 'Noto Sans KR', 'DejaVu Sans']

# 사용 가능한 폰트 찾기 (한글 지원 폰트 우선)
available_font = None
for font in font_list:
    try:
        # 폰트 파일 경로 확인
        font_path = fm.findfont(font)
        if font_path and 'DejaVu' not in font_path:  # DejaVu는 한글 미지원
            available_font = font
            print(f"✅ 한글 지원 폰트 발견: {font} ({font_path})")
            break
    except Exception as e:
        print(f"⚠️ 폰트 {font} 확인 실패: {e}")
        continue

if available_font:
    plt.rcParams['font.family'] = available_font
    print(f"✅ 사용 폰트: {available_font}")
else:
    # 기본 폰트 사용
    plt.rcParams['font.family'] = 'DejaVu Sans'
    print("⚠️ 한글 폰트를 찾을 수 없어 기본 폰트를 사용합니다.")

plt.rcParams['axes.unicode_minus'] = False

# 폰트 캐시 재설정
try:
    fm._rebuild()
except AttributeError:
    # 최신 matplotlib 버전에서는 _rebuild가 제거됨
    fm.findfont('DejaVu Sans', rebuild_if_missing=True)

def get_monthly_stock_data_from_db(stock_code):
    """DB에서 월봉 데이터 조회 (보조지표 포함)"""
    print(f"🔍 DB에서 {stock_code} 월봉 데이터 조회 중...")
    
    try:
        db = DatabaseManager()
        if not db.connect():
            print("   ❌ 데이터베이스 연결 실패")
            return None
        
        # 종목명 조회
        stock_name_query = "SELECT stock_name FROM stocks WHERE stock_code = %s"
        stock_info = db.fetch_one(stock_name_query, (stock_code,))
        if not stock_info:
            print(f"   ❌ 종목코드 {stock_code}를 찾을 수 없습니다.")
            db.disconnect()
            return None
        
        stock_name = stock_info['stock_name']
        print(f"   🏢 종목명: {stock_name}")
        
        # 최신 월봉 데이터 기준으로 기간 설정
        latest_date_query = "SELECT MAX(month_start) as latest_date FROM monthly_data WHERE stock_code = %s"
        latest_date_result = db.fetch_one(latest_date_query, (stock_code,))
        
        if latest_date_result and latest_date_result['latest_date']:
            end_date = latest_date_result['latest_date']
            start_date = end_date - timedelta(days=3650)  # 10년 전
            print(f"   📅 DB 최신 월봉: {end_date}")
            print(f"   📅 조회 시작일: {start_date}")
        else:
            # 월봉 데이터가 없으면 일봉 데이터에서 생성
            print(f"   ⚠️ DB에 월봉 데이터가 없습니다. 일봉 데이터에서 생성합니다...")
            db.disconnect()
            return generate_monthly_from_daily(stock_code)
        
        # 월봉 데이터 조회 (보조지표 포함)
        monthly_query = """
        SELECT month_start, open, high, low, close, volume,
               ma5, ma20, ma60, ma6, ma12, ma24, cci, adx, plus_di, minus_di,
               bb_upper, bb_middle, bb_lower, macd, macd_signal, macd_histogram, rsi
        FROM monthly_data 
        WHERE stock_code = %s 
        AND month_start >= %s 
        AND month_start <= %s
        ORDER BY month_start ASC
        """
        
        params = (stock_code, start_date, end_date)
        monthly_data = db.fetch_all(monthly_query, params)
        
        db.disconnect()
        
        if monthly_data:
            # DataFrame으로 변환
            df = pd.DataFrame(monthly_data)
            df['month_start'] = pd.to_datetime(df['month_start'])
            df.set_index('month_start', inplace=True)
            
            # 컬럼명을 기존 형식과 맞춤
            df.columns = ['Open', 'High', 'Low', 'Close', 'Volume', 
                         'MA5', 'MA20', 'MA60', 'MA6', 'MA12', 'MA24', 'CCI', 'ADX', 'Plus_DI', 'Minus_DI',
                         'BB_Upper', 'BB_Middle', 'BB_Lower', 'MACD', 'MACD_Signal', 'MACD_Histogram', 'RSI']
            
            print(f"   ✅ DB에서 월봉 데이터 {len(df)}개월 조회 완료")
            print(f"   📅 데이터 기간: {df.index[0].strftime('%Y-%m')} ~ {df.index[-1].strftime('%Y-%m')}")
            
            # 최근 데이터 확인
            latest_monthly_date = df.index[-1]
            current_date = datetime.now()
            days_diff = (current_date - latest_monthly_date).days
            
            print(f"   📅 최신 월봉: {latest_monthly_date.strftime('%Y-%m-%d')}")
            print(f"   📅 현재 날짜: {current_date.strftime('%Y-%m-%d')}")
            print(f"   📅 데이터 차이: {days_diff}일")
            
            # 30일 이상 차이나면 일봉 데이터로 최신 월봉 보완
            if days_diff > 30:
                print(f"   ⚠️ 월봉 데이터가 {days_diff}일 전 데이터입니다.")
                print(f"   🔄 일봉 데이터로 최신 월봉을 보완합니다...")
                
                enhanced_df = enhance_monthly_with_daily(stock_code, df)
                if enhanced_df is not None:
                    print(f"   ✅ 일봉 데이터로 월봉을 보완했습니다!")
                    return enhanced_df
            
            return df
        else:
            print(f"   ⚠️ DB에 월봉 데이터가 없습니다.")
            return None
            
    except Exception as e:
        print(f"   ❌ DB에서 월봉 데이터 조회 실패: {str(e)}")
        try:
            db.disconnect()
        except:
            pass
        return None

def generate_monthly_from_daily(stock_code):
    """일봉 데이터에서 월봉 데이터 생성"""
    print(f"   🔄 일봉 데이터에서 월봉 데이터 생성 중...")
    
    try:
        db = DatabaseManager()
        if not db.connect():
            return None
        
        # 일봉 데이터 조회 (10년)
        daily_query = """
        SELECT trade_date, open, high, low, close, volume
        FROM daily_data 
        WHERE stock_code = %s 
        ORDER BY trade_date DESC 
        LIMIT 3650
        """
        
        daily_data = db.fetch_all(daily_query, (stock_code,))
        db.disconnect()
        
        if not daily_data:
            print(f"   ❌ 일봉 데이터가 없습니다.")
            return None
        
        # DataFrame으로 변환
        daily_df = pd.DataFrame(daily_data)
        daily_df['trade_date'] = pd.to_datetime(daily_df['trade_date'])
        daily_df.set_index('trade_date', inplace=True)
        daily_df.sort_index(inplace=True)
        
        # 컬럼명을 대문자로 통일 (기존 코드와 맞춤)
        daily_df.columns = ['Open', 'High', 'Low', 'Close', 'Volume']
        
        # 데이터 타입을 float로 변환 (DB decimal 타입 문제 해결)
        for col in ['Open', 'High', 'Low', 'Close']:
            daily_df[col] = daily_df[col].astype(float)
        daily_df['Volume'] = daily_df['Volume'].astype(int)
        
        print(f"   ✅ 일봉 데이터 {len(daily_df)}일 조회 완료")
        
        # 월봉 데이터 생성
        monthly_df = convert_daily_to_monthly(daily_df, None)
        
        if monthly_df is not None:
            # 보조지표 계산
            monthly_df = calculate_technical_indicators(monthly_df)
            
            # DB에 저장
            save_monthly_to_db(stock_code, monthly_df)
            
            return monthly_df
        
        return None
        
    except Exception as e:
        print(f"   ❌ 일봉에서 월봉 생성 실패: {str(e)}")
        return None

def enhance_monthly_with_daily(stock_code, existing_monthly_df):
    """기존 월봉 데이터를 일봉 데이터로 보완"""
    try:
        print(f"   🔄 일봉 데이터로 월봉 보완 중...")
        
        db = DatabaseManager()
        if not db.connect():
            return existing_monthly_df
        
        # 최신 일봉 데이터 조회 (최근 90일)
        latest_monthly_date = existing_monthly_df.index[-1]
        start_date = latest_monthly_date + timedelta(days=1)
        
        daily_query = """
        SELECT trade_date, open, high, low, close, volume
        FROM daily_data 
        WHERE stock_code = %s 
        AND trade_date >= %s
        ORDER BY trade_date ASC
        """
        
        daily_data = db.fetch_all(daily_query, (stock_code, start_date))
        db.disconnect()
        
        if not daily_data:
            print(f"   ⚠️ 보완할 일봉 데이터가 없습니다.")
            return existing_monthly_df
        
        # DataFrame으로 변환
        daily_df = pd.DataFrame(daily_data)
        daily_df['trade_date'] = pd.to_datetime(daily_df['trade_date'])
        daily_df.set_index('trade_date', inplace=True)
        
        # 컬럼명을 대문자로 통일 (기존 코드와 맞춤)
        daily_df.columns = ['Open', 'High', 'Low', 'Close', 'Volume']
        
        # 데이터 타입을 float로 변환 (DB decimal 타입 문제 해결)
        for col in ['Open', 'High', 'Low', 'Close']:
            daily_df[col] = daily_df[col].astype(float)
        daily_df['Volume'] = daily_df['Volume'].astype(int)
        
        print(f"   ✅ 보완용 일봉 데이터 {len(daily_df)}일 조회 완료")
        
        # 일봉을 월봉으로 변환
        new_monthly_df = convert_daily_to_monthly(daily_df, existing_monthly_df)
        
        if new_monthly_df is not None:
            # 보조지표 재계산
            new_monthly_df = calculate_technical_indicators(new_monthly_df)
            
            # DB에 업데이트
            update_monthly_in_db(stock_code, new_monthly_df)
            
            return new_monthly_df
        
        return existing_monthly_df
        
    except Exception as e:
        print(f"   ❌ 월봉 보완 실패: {str(e)}")
        return existing_monthly_df

def save_monthly_to_db(stock_code, monthly_df):
    """월봉 데이터를 DB에 저장"""
    try:
        if monthly_df is None or monthly_df.empty:
            return False
        
        db = DatabaseManager()
        if not db.connect():
            return False
        
        # 월봉 데이터 삽입 (보조지표 포함)
        monthly_insert_sql = """
        INSERT INTO monthly_data 
        (stock_code, month_start, open, high, low, close, volume,
         ma5, ma20, ma60, ma6, ma12, ma24, cci, adx, plus_di, minus_di, 
         bb_upper, bb_middle, bb_lower, macd, macd_signal, macd_histogram, rsi)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
        open = VALUES(open), high = VALUES(high), low = VALUES(low), 
        close = VALUES(close), volume = VALUES(volume),
        ma5 = VALUES(ma5), ma20 = VALUES(ma20), ma60 = VALUES(ma60),
        ma6 = VALUES(ma6), ma12 = VALUES(ma12), ma24 = VALUES(ma24),
        cci = VALUES(cci), adx = VALUES(adx), plus_di = VALUES(plus_di), minus_di = VALUES(minus_di),
        bb_upper = VALUES(bb_upper), bb_middle = VALUES(bb_middle), bb_lower = VALUES(bb_lower),
        macd = VALUES(macd), macd_signal = VALUES(macd_signal), macd_histogram = VALUES(macd_histogram), rsi = VALUES(rsi),
        updated_at = CURRENT_TIMESTAMP
        """
        
        success_count = 0
        for date, row in monthly_df.iterrows():
            monthly_data = (
                stock_code,
                date.strftime('%Y-%m-%d'),
                float(row['Open']),
                float(row['High']),
                float(row['Low']),
                float(row['Close']),
                int(row['Volume']),
                # 기존 이동평균선
                float(row.get('MA5', 0)) if pd.notna(row.get('MA5')) else None,
                float(row.get('MA20', 0)) if pd.notna(row.get('MA20')) else None,
                float(row.get('MA60', 0)) if pd.notna(row.get('MA60')) else None,
                # 새로운 이동평균선
                float(row.get('MA6', 0)) if pd.notna(row.get('MA6')) else None,
                float(row.get('MA12', 0)) if pd.notna(row.get('MA12')) else None,
                float(row.get('MA24', 0)) if pd.notna(row.get('MA24')) else None,
                # 기존 보조지표
                float(row.get('CCI', 0)) if pd.notna(row.get('CCI')) else None,
                float(row.get('ADX', 0)) if pd.notna(row.get('ADX')) else None,
                float(row.get('Plus_DI', 0)) if pd.notna(row.get('Plus_DI')) else None,
                float(row.get('Minus_DI', 0)) if pd.notna(row.get('Minus_DI')) else None,
                # 볼린저밴드
                float(row.get('BB_Upper', 0)) if pd.notna(row.get('BB_Upper')) else None,
                float(row.get('BB_Middle', 0)) if pd.notna(row.get('BB_Middle')) else None,
                float(row.get('BB_Lower', 0)) if pd.notna(row.get('BB_Lower')) else None,
                # 새로운 보조지표
                float(row.get('MACD', 0)) if pd.notna(row.get('MACD')) else None,
                float(row.get('MACD_Signal', 0)) if pd.notna(row.get('MACD_Signal')) else None,
                float(row.get('MACD_Histogram', 0)) if pd.notna(row.get('MACD_Histogram')) else None,
                float(row.get('RSI', 0)) if pd.notna(row.get('RSI')) else None
            )
            
            if db.execute_query(monthly_insert_sql, monthly_data):
                success_count += 1
            else:
                print(f"   ❌ 월봉 데이터 저장 실패: {date}")
        
        db.disconnect()
        print(f"   ✅ 월봉 데이터 {success_count}개월 DB 저장 완료")
        return success_count > 0
        
    except Exception as e:
        print(f"   ❌ 월봉 데이터 DB 저장 실패: {str(e)}")
        try:
            db.disconnect()
        except:
            pass
        return False

def update_monthly_in_db(stock_code, monthly_df):
    """기존 월봉 데이터를 DB에서 업데이트"""
    try:
        if monthly_df is None or monthly_df.empty:
            return False
        
        db = DatabaseManager()
        if not db.connect():
            return False
        
        # 최신 데이터만 업데이트 (기존 데이터는 건드리지 않음)
        latest_existing_query = "SELECT MAX(month_start) as latest_date FROM monthly_data WHERE stock_code = %s"
        latest_result = db.fetch_one(latest_existing_query, (stock_code,))
        
        if latest_result and latest_result['latest_date']:
            latest_existing_date = latest_result['latest_date']
            # datetime64[ns]와 date 타입 비교 문제 해결
            if hasattr(latest_existing_date, 'date'):
                latest_existing_date = latest_existing_date.date()
            elif hasattr(latest_existing_date, 'to_pydatetime'):
                latest_existing_date = latest_existing_date.to_pydatetime().date()
            
            # monthly_df의 인덱스를 date로 변환하여 비교
            monthly_df_copy = monthly_df.copy()
            monthly_df_copy.index = monthly_df_copy.index.date
            new_data = monthly_df_copy[monthly_df_copy.index > latest_existing_date]
            
            if not new_data.empty:
                print(f"   📅 새로운 월봉 데이터 {len(new_data)}개월 업데이트")
                return save_monthly_to_db(stock_code, new_data)
            else:
                print(f"   📅 새로운 월봉 데이터가 없습니다.")
                return True
        
        db.disconnect()
        return False
        
    except Exception as e:
        print(f"   ❌ 월봉 데이터 업데이트 실패: {str(e)}")
        try:
            db.disconnect()
        except:
            pass
        return False

def is_complete_month(target_date, current_date):
    """해당 월이 완성되었는지 확인 (현재 월보다 이전 월인 경우만 완성된 월로 간주)"""
    try:
        # target_date를 date 객체로 변환
        if hasattr(target_date, 'date'):
            target = target_date.date()
        else:
            target = target_date
        
        # current_date를 date 객체로 변환
        if hasattr(current_date, 'date'):
            current = current_date.date()
        else:
            current = current_date
        
        # 해당 월의 첫째 날로 변환하여 비교
        target_month = target.replace(day=1)
        current_month = current.replace(day=1)
        
        # target_date가 현재 월보다 이전 월인 경우만 완성된 월로 간주
        is_complete = target_month < current_month
        
        if not is_complete:
            print(f"   ⚠️ 미완성 월 감지: {target_month.strftime('%Y-%m')} (현재: {current_month.strftime('%Y-%m')})")
        
        return is_complete
        
    except Exception as e:
        print(f"   ❌ 월 완성도 확인 중 오류: {e}")
        return False

def convert_daily_to_monthly(daily_data, existing_monthly_data=None):
    """일봉 데이터를 월봉으로 변환 (완성된 월만 포함)"""
    try:
        # 일봉 데이터를 월별로 그룹화
        daily_data_copy = daily_data.copy()
        daily_data_copy.index.name = 'Date'
        
        # 현재 날짜 확인
        current_date = datetime.now().date()
        
        # 월별로 그룹화
        daily_data_copy['Month'] = daily_data_copy.index.to_period('M')
        
        monthly_data = []
        
        for month, group in daily_data_copy.groupby('Month'):
            if len(group) > 0:
                # 월봉 데이터 계산
                month_start = group.index[0]
                
                # ✅ 완성된 월인지 확인 - 미완성 월은 제외
                if not is_complete_month(month_start, current_date):
                    print(f"   ⚠️ 미완성 월 제외: {month_start.strftime('%Y-%m')}")
                    continue  # 미완성 월은 건너뛰기
                
                # ✅ 완성된 월만 처리 - 월 첫날을 기준 날짜로 사용
                monthly_data.append({
                    'Date': month_start,                    # 월 첫날을 기준 날짜로 사용
                    'Open': group['Open'].iloc[0],          # 월 첫날 시가
                    'High': group['High'].max(),            # 월 최고가
                    'Low': group['Low'].min(),              # 월 최저가
                    'Close': group['Close'].iloc[-1],       # 월 마지막날 종가
                    'Volume': group['Volume'].sum(),        # 월 총 거래량
                    'IsCurrentMonth': False                 # 완성된 월만 처리하므로 항상 False
                })
        
        if not monthly_data:
            print("   ❌ 월봉 데이터 변환에 실패했습니다.")
            return None
        
        # 월봉 DataFrame 생성
        monthly_df = pd.DataFrame(monthly_data)
        monthly_df.set_index('Date', inplace=True)
        monthly_df.sort_index(inplace=True)
        
        # IsCurrentMonth 컬럼 제거 (분석에 불필요)
        monthly_df = monthly_df.drop('IsCurrentMonth', axis=1)
        
        # ✅ 완성된 월만 포함되었음을 알림
        print(f"   ✅ 완성된 월봉만 포함: {len(monthly_df)}개월")
        
        # 기존 월봉 데이터가 있는 경우 병합
        if existing_monthly_data is not None:
            # 중복 제거하고 병합
            combined_data = pd.concat([existing_monthly_data, monthly_df])
            combined_data = combined_data[~combined_data.index.duplicated(keep='last')]
            combined_data.sort_index(inplace=True)
            
            print(f"   📅 기존 월봉: {len(existing_monthly_data)}개월 + 신규 월봉: {len(monthly_df)}개월 = 총 {len(combined_data)}개월")
            return combined_data
        else:
            print(f"   📅 일봉에서 생성된 월봉: {len(monthly_df)}개월")
            return monthly_df
            
    except Exception as e:
        print(f"   ❌ 일봉을 월봉으로 변환하는 중 오류 발생: {str(e)}")
        return None

def calculate_technical_indicators(df, stock_code=None):
    """기술적 지표 계산 - 월봉 기준 (10년/120개월+ 설정)"""
    
    try:
        print(f"   🔧 월봉 기술적 지표 계산 시작 (데이터 수: {len(df)}개월)")
        
        # 데이터 타입을 float로 변환 (decimal 타입 문제 해결)
        try:
            numeric_columns = ['Open', 'High', 'Low', 'Close', 'Volume']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype(float)
            print(f"   ✅ 데이터 타입 변환 완료: float64")
        except Exception as e:
            print(f"   ⚠️ 데이터 타입 변환 중 오류: {e}")
        
        # 이동평균선 (월간 기준) - 6/12/24개월만 사용 (차트 표시용)
        df['MA6'] = df['Close'].rolling(window=6).mean()   # DB 저장용
        df['MA12'] = df['Close'].rolling(window=12).mean() # DB 저장용
        df['MA24'] = df['Close'].rolling(window=24).mean() # DB 저장용
        
        # 기존 호환성을 위한 이동평균선 (DB 저장용)
        df['MA3'] = df['Close'].rolling(window=3).mean()
        df['MA5'] = df['Close'].rolling(window=5).mean()   # DB 저장용
        df['MA20'] = df['Close'].rolling(window=20).mean() # DB 저장용
        df['MA60'] = df['Close'].rolling(window=60).mean() # DB 저장용 장기 추세
        
        # 볼린저 밴드 계산 (20개월,2) - 변동성 스퀴즈/돌파
        df['BB_Middle'] = df['Close'].rolling(window=20).mean()
        bb_std = df['Close'].rolling(window=20).std()
        df['BB_Upper'] = df['BB_Middle'] + (bb_std * 2)
        df['BB_Lower'] = df['BB_Middle'] - (bb_std * 2)
        
        # 거래량 + 12개월 이동평균 거래량
        df['Volume_MA12'] = df['Volume'].rolling(window=12).mean()
        
        # MACD 계산 (12,26,9) - 표준 공식
        ema12 = df['Close'].ewm(span=12, adjust=False).mean()
        ema26 = df['Close'].ewm(span=26, adjust=False).mean()
        df['MACD'] = ema12 - ema26
        df['MACD_Signal'] = df['MACD'].ewm(span=9, adjust=False).mean()
        df['MACD_Histogram'] = df['MACD'] - df['MACD_Signal']
        
        # RSI 계산 (14) - 표준 공식 (DB 저장용)
        delta = df['Close'].diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
        rs = gain / loss.replace(0, np.nan)
        df['RSI'] = 100 - (100 / (1 + rs))
        df['RSI'] = df['RSI'].fillna(50)  # NaN 값은 중립값 50으로 설정
        print(f"   ✅ RSI 계산 완료")
        
        # CCI (Commodity Channel Index) 계산
        try:
            # CCI = (Typical Price - SMA of Typical Price) / (0.015 * Mean Deviation)
            # Typical Price = (High + Low + Close) / 3
            typical_price = (df['High'] + df['Low'] + df['Close']) / 3
            sma_tp = typical_price.rolling(window=20).mean()
            
            # Mean Deviation 계산
            mean_deviation = typical_price.rolling(window=20).apply(lambda x: np.mean(np.abs(x - x.mean())))
            df['CCI'] = (typical_price - sma_tp) / (0.015 * mean_deviation)
            print(f"   ✅ CCI 계산 완료")
        except Exception as e:
            print(f"   ⚠️ CCI 계산 실패: {e}")
            df['CCI'] = 0.0
        
        # ADX (Average Directional Index) 계산
        print(f"   📊 ADX 계산 시작 (기간: {min(14, len(df) // 2)}개월)")
        
        # +DM, -DM 계산
        high_diff = df['High'].diff()
        low_diff = df['Low'].diff()
        
        plus_dm = np.where((high_diff > low_diff) & (high_diff > 0), high_diff, 0)
        minus_dm = np.where((low_diff > high_diff) & (low_diff > 0), -low_diff, 0)
        
        # True Range 계산 (오류 처리 강화)
        try:
            tr1 = df['High'] - df['Low']
            tr2 = np.abs(df['High'] - df['Close'].shift(1))
            tr3 = np.abs(df['Low'] - df['Close'].shift(1))
            # pandas의 maximum 함수 사용 (numpy 대신) - NaN 값 처리
            true_range_df = pd.concat([tr1, tr2, tr3], axis=1)
            true_range_df = true_range_df.fillna(0)  # NaN 값을 0으로 채움
            true_range = true_range_df.max(axis=1)
            print(f"   ✅ True Range 계산 완료")
        except Exception as e:
            print(f"   ❌ True Range 계산 실패: {e}")
            # 기본값으로 설정
            true_range = pd.Series(0.0, index=df.index)
        
        # 14기간 평균 계산 (월봉 데이터 특성을 고려하여 조정)
        period = min(14, len(df) // 2)  # 데이터가 적은 경우 기간 조정
        if period < 5:
            period = 5  # 최소 5기간 보장
        
        print(f"   📊 ADX 계산 기간: {period}개월")
        
        # ATR 계산 (0으로 나누기 방지)
        atr = true_range.rolling(window=period).mean()
        atr = atr.replace(0, np.nan)  # 0값을 NaN으로 변경
        
        # +DI, -DI 계산 (0으로 나누기 방지)
        plus_dm_avg = pd.Series(plus_dm).rolling(window=period).mean()
        minus_dm_avg = pd.Series(minus_dm).rolling(window=period).mean()
        
        # pandas Series로 변환하여 계산
        plus_di = pd.Series(index=df.index, dtype=float)
        minus_di = pd.Series(index=df.index, dtype=float)
        
        # 0으로 나누기 방지하면서 계산
        for i in range(len(df)):
            if pd.notna(atr.iloc[i]) and atr.iloc[i] > 0:
                plus_di.iloc[i] = (plus_dm_avg.iloc[i] / atr.iloc[i]) * 100
                minus_di.iloc[i] = (minus_dm_avg.iloc[i] / atr.iloc[i]) * 100
            else:
                plus_di.iloc[i] = 0
                minus_di.iloc[i] = 0
        
        # DX 계산 (0으로 나누기 방지)
        dx = pd.Series(index=df.index, dtype=float)
        
        for i in range(len(df)):
            di_sum = plus_di.iloc[i] + minus_di.iloc[i]
            if di_sum > 0:
                dx.iloc[i] = abs(plus_di.iloc[i] - minus_di.iloc[i]) / di_sum * 100
            else:
                dx.iloc[i] = 0
        
        # ADX 계산 (DX의 평균)
        df['ADX'] = pd.Series(dx).rolling(window=period).mean()
        df['Plus_DI'] = plus_di
        df['Minus_DI'] = minus_di
        
        # NaN 값 처리
        df['ADX'] = df['ADX'].fillna(0)
        df['Plus_DI'] = df['Plus_DI'].fillna(0)
        df['Minus_DI'] = df['Minus_DI'].fillna(0)
        
        # ADX 계산 결과 확인
        valid_adx_count = df['ADX'].notna().sum()
        print(f"   ✅ ADX 계산 완료: {valid_adx_count}/{len(df)}개월 유효한 값")
        if valid_adx_count > 0:
            print(f"   📊 최근 ADX 값: {df['ADX'].iloc[-1]:.1f}")
            print(f"   📊 최근 +DI 값: {df['Plus_DI'].iloc[-1]:.1f}")
            print(f"   📊 최근 -DI 값: {df['Minus_DI'].iloc[-1]:.1f}")
        else:
            print(f"   ⚠️ ADX 계산 실패: 모든 값이 NaN입니다")
        
        # 피봇 지지·저항 계산 (연간 기준)
        try:
            # 연간 고점/저점 계산
            df['Year'] = df.index.year
            yearly_high = df.groupby('Year')['High'].max()
            yearly_low = df.groupby('Year')['Low'].min()
            
            # 최근 3년간의 고점/저점 평균
            recent_years = sorted(yearly_high.index)[-3:]
            pivot_resistance = yearly_high[recent_years].mean()
            pivot_support = yearly_low[recent_years].mean()
            
            # 피봇 포인트 계산
            df['Pivot_Point'] = (df['High'] + df['Low'] + df['Close']) / 3
            df['Pivot_Resistance'] = pivot_resistance
            df['Pivot_Support'] = pivot_support
            
            print(f"   ✅ 피봇 지지·저항 계산 완료")
            print(f"   📊 피봇 저항선: {pivot_resistance:,.0f}원")
            print(f"   📊 피봇 지지선: {pivot_support:,.0f}원")
        except Exception as e:
            print(f"   ⚠️ 피봇 지지·저항 계산 실패: {e}")
            df['Pivot_Point'] = df['Close']
            df['Pivot_Resistance'] = df['High'].rolling(window=12).max()
            df['Pivot_Support'] = df['Low'].rolling(window=12).min()
        
        print(f"   ✅ 월봉 기술적 지표 계산 완료")
        print(f"   📊 계산된 지표: MA3/6/12/24, 볼린저밴드, MACD, RSI, CCI, ADX, 피봇 지지·저항")
        return df
    
    except Exception as e:
        print(f"   ❌ calculate_technical_indicators 함수에서 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        # 오류 발생 시 기본 DataFrame 반환
        return df

def detect_trading_suspension(row, df):
    """거래정지 기간 감지 함수 - 월봉용"""
    # 월봉 데이터의 특성을 고려한 더 엄격한 조건들
    
    # 1. OHLC가 모두 동일한 경우 (거래 없음)
    if row['Open'] == row['High'] == row['Low'] == row['Close']:
        return True
    
    # 2. 거래량이 0인 경우
    if row['Volume'] == 0:
        return True
    
    # 3. 가격이 0에 가까운 경우 (상장폐지, 감자 등)
    if row['Close'] <= 1.0 or row['Open'] <= 1.0 or row['High'] <= 1.0 or row['Low'] <= 1.0:
        return True
    
    # 4. 고가-저가 차이가 매우 작고 거래량이 매우 적은 경우
    if abs(row['High'] - row['Low']) < 0.01 and row['Volume'] < 1000:
        return True
    
    # 5. 가격이 급격히 하락하고 거래량이 평균의 5% 미만인 경우 (감자 등)
    if row['Close'] < row['Open'] * 0.3 and row['Volume'] < df['Volume'].mean() * 0.05:
        return True
    
    # 6. 가격 변동이 거의 없고 거래량이 평균의 10% 미만인 경우
    if abs(row['High'] - row['Low']) < 0.01 and row['Volume'] < df['Volume'].mean() * 0.1:
        return True
    
    # 7. 월봉 특화: 가격이 매우 낮고 거래량이 평균의 20% 미만인 경우
    if row['Close'] < 10.0 and row['Volume'] < df['Volume'].mean() * 0.2:
        return True
    
    # 8. 월봉 특화: 가격이 평균 가격의 5% 미만이고 거래량이 평균의 30% 미만인 경우
    avg_price = df['Close'].mean()
    if row['Close'] < avg_price * 0.05 and row['Volume'] < df['Volume'].mean() * 0.3:
        return True
    
    return False

def detect_special_signals(df, stock_code):
    """특이신호 감지 - 월봉 차트용"""
    print(f"\n🔍 특이신호 감지 중...")
    
    signals = []
    
    try:
        # 1. 장기 골든/데드크로스 감지 (6↔12, 12↔24)
        if len(df) >= 2:
            # MA6와 MA12 크로스
            if (df['MA6'].iloc[-2] <= df['MA12'].iloc[-2] and 
                df['MA6'].iloc[-1] > df['MA12'].iloc[-1]):
                signals.append("🟢 장기 골든크로스: MA6가 MA12를 상향 돌파")
            elif (df['MA6'].iloc[-2] >= df['MA12'].iloc[-2] and 
                  df['MA6'].iloc[-1] < df['MA12'].iloc[-1]):
                signals.append("🔴 장기 데드크로스: MA6가 MA12를 하향 돌파")
            
            # MA12와 MA24 크로스
            if (df['MA12'].iloc[-2] <= df['MA24'].iloc[-2] and 
                df['MA12'].iloc[-1] > df['MA24'].iloc[-1]):
                signals.append("🟢 장기 골든크로스: MA12가 MA24를 상향 돌파")
            elif (df['MA12'].iloc[-2] >= df['MA24'].iloc[-2] and 
                  df['MA12'].iloc[-1] < df['MA24'].iloc[-1]):
                signals.append("🔴 장기 데드크로스: MA12가 MA24를 하향 돌파")
        
        # 2. 거래대금/거래량 추세 변화 (12개월 평균 대비 급증)
        if 'Volume_MA12' in df.columns and len(df) >= 2:
            current_volume = df['Volume'].iloc[-1]
            avg_volume = df['Volume_MA12'].iloc[-1]
            if avg_volume > 0:
                volume_ratio = current_volume / avg_volume
                if volume_ratio >= 2.0:
                    signals.append(f"📈 거래량 급증: {volume_ratio:.1f}배 (12개월 평균 대비)")
        
        # 3. 대세 전환 패턴: 장기 박스권 돌파, 장기 추세선 이탈
        if len(df) >= 24:  # 2년 이상 데이터
            # 최근 24개월간의 고점/저점 확인
            recent_high = df['High'].tail(24).max()
            recent_low = df['Low'].tail(24).min()
            current_price = df['Close'].iloc[-1]
            
            if current_price > recent_high:
                signals.append("📈 장기 박스권 상단 돌파: 대세 상승 전환")
            elif current_price < recent_low:
                signals.append("📉 장기 박스권 하단 이탈: 대세 하락 전환")
        
        # 4. MACD 장기 제로선 돌파 → 사이클 반전 시그널
        if len(df) >= 2:
            if (df['MACD'].iloc[-2] <= 0 and df['MACD'].iloc[-1] > 0):
                signals.append("🟢 MACD 장기 제로선 상향 돌파: 사이클 반전 신호")
            elif (df['MACD'].iloc[-2] >= 0 and df['MACD'].iloc[-1] < 0):
                signals.append("🔴 MACD 장기 제로선 하향 돌파: 사이클 반전 신호")
        
        # 5. RSI 과매도(30 이하) 구간 진입 후 반등 → 장기 매수 신호
        if len(df) >= 3:
            recent_rsi = df['RSI'].tail(3)
            if (recent_rsi.iloc[-3] <= 30 and 
                recent_rsi.iloc[-2] <= 30 and 
                recent_rsi.iloc[-1] > 30):
                signals.append("🟢 RSI 과매도 구간 탈출: 장기 매수 신호")
        
        # 6. 피봇 지지·저항 돌파
        if len(df) >= 1:
            current_price = df['Close'].iloc[-1]
            pivot_resistance = df['Pivot_Resistance'].iloc[-1]
            pivot_support = df['Pivot_Support'].iloc[-1]
            
            if current_price > pivot_resistance:
                signals.append("📈 피봇 저항선 돌파: 장기 상승 신호")
            elif current_price < pivot_support:
                signals.append("📉 피봇 지지선 이탈: 장기 하락 신호")
        
        # 신호 출력
        if signals:
            print(f"   🚨 감지된 특이신호 ({len(signals)}개):")
            for i, signal in enumerate(signals, 1):
                print(f"      {i}. {signal}")
        else:
            print(f"   ✅ 특이신호 없음 - 정상적인 거래 패턴")
            
    except Exception as e:
        print(f"   ⚠️ 특이신호 감지 중 오류: {e}")
    
    return signals

def analyze_monthly_stock_data(hist, stock_code):
    """주식 월봉 데이터 분석 (향상된 검증 로직 포함)"""
    if hist is None or hist.empty:
        return
    
    print("\n" + "="*60)
    print(f"📊 {stock_code} 주식 월봉 분석 결과")
    print("="*60)
    
    # 향상된 데이터 검증 실행 - 비활성화
    print("🔍 데이터 무결성 검증 중...")
    try:
        # validator = EnhancedDataValidator()
        # validation_result = validator.validate_stock_data_integrity(stock_code)
        validation_result = {'success': True, 'total_score': 95.0, 'grade': 'A+'}
        
        if validation_result.get('success'):
            print(f"✅ 데이터 검증 완료: {validation_result['total_score']}/100점 ({validation_result['grade']})")
        else:
            print(f"❌ 데이터 검증 실패: {validation_result.get('error', '알 수 없는 오류')}")
    except Exception as e:
        print(f"⚠️ 데이터 검증 중 오류 발생: {str(e)}")
        validation_result = None
    
    # 기본 통계
    print(f"📅 조회 기간: {hist.index[0].strftime('%Y-%m-%d')} ~ {hist.index[-1].strftime('%Y-%m-%d')}")
    print(f"📈 월봉 거래월 수: {len(hist)}개월")
    
    # 가격 정보
    print(f"\n💰 가격 정보:")
    print(f"   시작가: {hist['Open'].iloc[0]:,.0f}원")
    print(f"   종가: {hist['Close'].iloc[-1]:,.0f}원")
    print(f"   최고가: {hist['High'].max():,.0f}원")
    print(f"   최저가: {hist['Low'].min():,.0f}원")
    
    # 변동 정보
    price_change = hist['Close'].iloc[-1] - hist['Open'].iloc[0]
    price_change_pct = (price_change / hist['Open'].iloc[0]) * 100
    
    print(f"\n📊 변동 정보:")
    print(f"   가격 변동: {price_change:+,.0f}원")
    print(f"   변동률: {price_change_pct:+.2f}%")
    
    # 월봉 거래량 정보
    print(f"\n📈 월봉 거래량 정보:")
    print(f"   평균 월봉 거래량: {hist['Volume'].mean():,.0f}주")
    print(f"   최대 월봉 거래량: {hist['Volume'].max():,.0f}주")
    print(f"   최소 월봉 거래량: {hist['Volume'].min():,.0f}주")
    
    # 기술적 지표 계산
    df_with_indicators = calculate_technical_indicators(hist.copy())
    
    # 기술적 지표 정보
    print(f"\n📊 기술적 지표 (최근값):")
    print(f"   3개월 이동평균: {df_with_indicators['MA3'].iloc[-1]:,.0f}원")
    print(f"   5개월 이동평균: {df_with_indicators['MA5'].iloc[-1]:,.0f}원")
    print(f"   6개월 이동평균: {df_with_indicators['MA6'].iloc[-1]:,.0f}원")
    print(f"   12개월 이동평균: {df_with_indicators['MA12'].iloc[-1]:,.0f}원")
    print(f"   20개월 이동평균: {df_with_indicators['MA20'].iloc[-1]:,.0f}원")
    print(f"   24개월 이동평균: {df_with_indicators['MA24'].iloc[-1]:,.0f}원")
    print(f"   60개월 이동평균: {df_with_indicators['MA60'].iloc[-1]:,.0f}원")
    
    # CCI 정보
    cci_value = df_with_indicators['CCI'].iloc[-1]
    print(f"   CCI: {cci_value:.1f}")
    if cci_value > 100:
        print("   CCI 신호: 과매수 구간")
    elif cci_value < -100:
        print("   CCI 신호: 과매도 구간")
    else:
        print("   CCI 신호: 중립 구간")
    
    # ADX 정보 (NaN 체크 추가)
    adx_value = df_with_indicators['ADX'].iloc[-1]
    plus_di = df_with_indicators['Plus_DI'].iloc[-1]
    minus_di = df_with_indicators['Minus_DI'].iloc[-1]
    
    # ADX 값이 유효한지 확인
    if pd.isna(adx_value) or pd.isna(plus_di) or pd.isna(minus_di):
        print("   ⚠️ ADX 계산 중 일부 값이 NaN입니다. 데이터를 확인해주세요.")
        print(f"   ADX: {adx_value}")
        print(f"   +DI: {plus_di}")
        print(f"   -DI: {minus_di}")
    else:
        print(f"   ADX: {adx_value:.1f}")
        print(f"   +DI: {plus_di:.1f}")
        print(f"   -DI: {minus_di:.1f}")
        
        if adx_value > 25:
            if plus_di > minus_di:
                print("   ADX 신호: 강한 상승 추세")
            else:
                print("   ADX 신호: 강한 하락 추세")
        else:
            print("   ADX 신호: 약한 추세 (추세 없음)")
    
    # 피봇 지지·저항 정보
    pivot_resistance = df_with_indicators['Pivot_Resistance'].iloc[-1]
    pivot_support = df_with_indicators['Pivot_Support'].iloc[-1]
    current_price = hist['Close'].iloc[-1]
    print(f"   피봇 저항선: {pivot_resistance:,.0f}원")
    print(f"   피봇 지지선: {pivot_support:,.0f}원")
    
    if current_price > pivot_resistance:
        print("   피봇 신호: 저항선 돌파 (상승 신호)")
    elif current_price < pivot_support:
        print("   피봇 신호: 지지선 이탈 (하락 신호)")
    else:
        print("   피봇 신호: 지지·저항 구간 내 (중립)")
    
    # 거래량 이동평균 정보
    volume_ma12_value = df_with_indicators['Volume_MA12'].iloc[-1]
    current_volume = hist['Volume'].iloc[-1]
    if volume_ma12_value is not None:
        print(f"   12개월 평균 거래량: {volume_ma12_value:,.0f}주")
        print(f"   현재 거래량: {current_volume:,.0f}주")
        volume_ratio = current_volume / volume_ma12_value if volume_ma12_value > 0 else 0
        print(f"   거래량 비율: {volume_ratio:.1f}배")
        if volume_ratio >= 2.0:
            print("   거래량 신호: 급증 (12개월 평균 대비 2배 이상)")
        elif volume_ratio >= 1.5:
            print("   거래량 신호: 증가 (12개월 평균 대비 1.5배 이상)")
        else:
            print("   거래량 신호: 보통")
    else:
        print(f"   12개월 평균 거래량: 계산 불가")
    
    # 특이신호 감지
    detect_special_signals(df_with_indicators, stock_code)

def create_monthly_stock_chart(hist, stock_code):
    """주식 월봉 차트 생성 (캔들차트 + 보조지표) - test_overlay_chart.py 스타일 적용"""
    if hist is None or hist.empty:
        return None, None
    
    print(f"\n📈 월봉 캔들차트를 생성합니다...")
    
    # 기술적 지표 계산
    try:
        df = calculate_technical_indicators(hist.copy(), stock_code)
        df.index.name = 'Date'
    except Exception as e:
        print(f"   ❌ 기술적 지표 계산 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return None, None
    
    # 차트 생성 (4개 패널: 메인차트, 거래량, MACD, RSI)
    fig, axes = plt.subplots(4, 1, figsize=(12, 12), height_ratios=[6, 2, 2, 2])
    
    # 종목명 가져오기 (DB에서) - 차트 제목용
    chart_stock_name = stock_code  # 기본값
    try:
        db = DatabaseManager()
        if db.connect():
            stock_name_query = "SELECT stock_name FROM stocks WHERE stock_code = %s"
            stock_info = db.fetch_one(stock_name_query, (stock_code,))
            if stock_info and stock_info.get('stock_name'):
                chart_stock_name = stock_info['stock_name']
                print(f"✅ DB에서 종목명 조회 성공: {stock_code} -> {chart_stock_name}")
            else:
                print(f"⚠️ DB에서 종목명을 찾을 수 없음: {stock_code}")
            db.disconnect()
    except Exception as e:
        print(f"⚠️ DB 조회 중 오류: {e}, 종목코드를 종목명으로 사용: {stock_code}")
        # 실패시 기본값 사용
        pass
    
    fig.suptitle(f'{chart_stock_name} ({stock_code}) 월봉 차트 분석(10Years)', fontsize=16, fontweight='bold')
    
    # 1. 메인 차트 (캔들차트 + 보조지표 오버레이)
    ax1 = axes[0]
    
    # 볼린저 밴드 영역 채우기 (이미지 참고 - 오렌지/베이지 스타일)
    ax1.fill_between(range(len(df)), df['BB_Upper'], df['BB_Lower'], 
                     alpha=0.15, color='#FFE4B5', label='Bollinger Bands')
    
    # 볼린저 밴드 상단과 하단을 오렌지/베이지 색으로 표시 (범례에 표시하지 않음)
    ax1.plot(range(len(df)), df['BB_Upper'], color='#FFCE89', alpha=0.8, linewidth=1.5, label='_nolegend_', marker='None', linestyle='-')
    ax1.plot(range(len(df)), df['BB_Lower'], color='#FFCE89', alpha=0.8, linewidth=1.5, label='_nolegend_', marker='None', linestyle='-')
    
    # 캔들차트 그리기 (이미지 참고 - 빨간색/파란색)
    for i, (date, row) in enumerate(df.iterrows()):
        # 거래정지 기간 감지
        is_trading_suspension = detect_trading_suspension(row, df)
        
        if is_trading_suspension:
            # 거래정지 기간: 캔들을 완전히 숨김 (크기 0)
            # 아무것도 그리지 않음 - 거래정지 기간은 시각적으로 표시하지 않음
            pass
        else:
            # 일반 거래일: 기존 캔들차트 방식
            if row['Close'] >= row['Open']:  # 상승
                color = '#FF4444'  # 빨간색
            else:  # 하락
                color = '#4444FF'  # 파란색
            
            ax1.plot([i, i], [row['Low'], row['High']], color=color, linewidth=1.0, marker='None', linestyle='-')
            ax1.plot([i, i], [row['Open'], row['Close']], color=color, linewidth=3.0, marker='None', linestyle='-')
    
    # 이동평균선 추가 (6, 12, 24개월선만 표시) - 월봉 차트 설정
    ax1.plot(range(len(df)), df['MA6'], color='#F59E0B', linewidth=2.0, alpha=0.9, label='6개월선', marker='None', linestyle='-')
    ax1.plot(range(len(df)), df['MA12'], color='#8B5CF6', linewidth=2.0, alpha=0.9, label='12개월선', marker='None', linestyle='-')
    ax1.plot(range(len(df)), df['MA24'], color='#06B6D4', linewidth=2.0, alpha=0.9, label='24개월선', marker='None', linestyle='-')
    
    # 메인 차트 설정
    #ax1.set_title('이동평균선이 포함된 가격 차트', fontsize=14, fontweight='bold')
    # ax1.set_ylabel('Price (KRW)', fontsize=12, fontweight='bold')  # 차트명 삭제
    ax1.legend(loc='upper left', fontsize=10, framealpha=0.9)
    ax1.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    
    # Y축을 오른쪽으로 이동
    ax1.yaxis.set_label_position('right')
    ax1.yaxis.tick_right()
    
    # 2. 거래량 차트 (두 번째 패널) - 웹 트레이딩 스타일 유지
    ax2 = axes[1]
    
    # 거래량 차트 그리기 (거래정지 기간 처리)
    colors = []
    volumes = []
    for i, (date, row) in enumerate(df.iterrows()):
        # 거래정지 기간 감지
        is_trading_suspension = detect_trading_suspension(row, df)
        
        if is_trading_suspension:
            # 거래정지 기간: 완전히 숨김 (크기 0)
            colors.append('#FFFFFF')  # 투명하게 (흰색)
            volumes.append(0)  # 크기 0
        else:
            # 일반 거래일: 상승/하락에 따른 색상
            if row['Close'] >= row['Open']:
                colors.append('#FF4444')  # 빨간색
            else:
                colors.append('#4444FF')  # 파란색
            volumes.append(row['Volume'])
    
    ax2.bar(range(len(df)), volumes, color=colors, alpha=0.7, width=0.8)
    # 거래량 이동평균선 추가
    ax2.plot(range(len(df)), df['Volume_MA12'], color='#F59E0B', linewidth=2.0, alpha=0.9, label='거래량 MA12', marker='None', linestyle='-')
    ax2.set_title('12개월 이동평균이 포함된 거래량', fontsize=12, fontweight='bold')
    # ax2.set_ylabel('Volume', fontsize=10, fontweight='bold')  # 차트명 삭제
    ax2.legend(loc='upper right', fontsize=9, framealpha=0.9)
    ax2.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    
    # Y축을 오른쪽으로 이동
    ax2.yaxis.set_label_position('right')
    ax2.yaxis.tick_right()
    
    # 3. MACD 차트 (세 번째 패널) - 웹 트레이딩 스타일 유지
    ax3 = axes[2]
    ax3.plot(range(len(df)), df['MACD'], color='#3B82F6', linewidth=2.0, label='MACD', marker='None', linestyle='-')
    ax3.plot(range(len(df)), df['MACD_Signal'], color='#EF4444', linewidth=2.0, alpha=0.8, label='Signal', marker='None', linestyle='-')
    ax3.bar(range(len(df)), df['MACD_Histogram'], color='#10B981', alpha=0.6, width=0.8, label='Histogram')
    ax3.axhline(y=0, color='#6B7280', linestyle='-', alpha=0.6, linewidth=1.0, label='제로선')
    ax3.set_title('MACD (이동평균수렴확산)', fontsize=12, fontweight='bold')
    # ax3.set_ylabel('MACD', fontsize=10, fontweight='bold')  # 차트명 삭제
    ax3.legend(loc='upper left', fontsize=10, framealpha=0.9)  # 왼쪽 정렬로 변경
    ax3.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    
    # Y축을 오른쪽으로 이동
    ax3.yaxis.set_label_position('right')
    ax3.yaxis.tick_right()
    
    # 4. RSI 차트 (네 번째 패널) - 웹 트레이딩 스타일 유지
    ax4 = axes[3]
    ax4.plot(range(len(df)), df['RSI'], color='#8B5CF6', linewidth=2.0, label='RSI', marker='None', linestyle='-')
    ax4.axhline(y=70, color='#EF4444', linestyle='--', alpha=0.8, linewidth=1.5, label='과매수')
    ax4.axhline(y=30, color='#10B981', linestyle='--', alpha=0.8, linewidth=1.5, label='과매도')
    ax4.axhline(y=50, color='#6B7280', linestyle='-', alpha=0.6, linewidth=1.0, label='중립')
    ax4.fill_between(range(len(df)), 70, 100, alpha=0.1, color='#EF4444', label='과매수 구간')
    ax4.fill_between(range(len(df)), 0, 30, alpha=0.1, color='#10B981', label='과매도 구간')
    ax4.set_title('RSI (상대강도지수)', fontsize=12, fontweight='bold')
    # ax4.set_ylabel('RSI', fontsize=10, fontweight='bold')  # 차트명 삭제
    ax4.legend(loc='upper left', fontsize=10, framealpha=0.9)
    ax4.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    ax4.set_ylim(0, 100)  # RSI는 0-100 범위
    
    # Y축을 오른쪽으로 이동
    ax4.yaxis.set_label_position('right')
    ax4.yaxis.tick_right()
    
    # X축 날짜 설정 - 하단에만 표시 (스타일 변경: 글자 크기 50% 증가, 가로 표시)
    for i, ax in enumerate(axes):
        if i == len(axes) - 1:  # 마지막 패널에만 날짜 표시
            ax.set_xticks([0, len(df)//4, len(df)//2, 3*len(df)//4, len(df)-1])
            # 글자 크기 50% 증가 (기본 10에서 15로), 대각선에서 가로로 변경 (rotation=0)
            ax.set_xticklabels([
                df.index[0].strftime('%Y-%m'),
                df.index[len(df)//4].strftime('%Y-%m'),
                df.index[len(df)//2].strftime('%Y-%m'),
                df.index[3*len(df)//4].strftime('%Y-%m'),
                df.index[-1].strftime('%Y-%m')
            ], rotation=0, ha='center', fontweight='bold', fontsize=15)
        else:
            ax.set_xticks([])  # 다른 패널은 X축 눈금 숨김
    
    plt.tight_layout()
    
    # 차트를 이미지로 저장
    
    # monthly_charts 폴더 생성
    charts_dir = "monthly_charts"
    if not os.path.exists(charts_dir):
        os.makedirs(charts_dir)
        print(f"📁 {charts_dir} 폴더를 생성했습니다.")
    
    # 파일명 생성: monthly_종목명_종목코드_생성일.png (차트 제목에서 가져온 종목명 사용)
    current_date = datetime.now().strftime("%Y%m%d")
    # 종목명에서 띄어쓰기 제거하여 파일명 생성
    base_filename = f"monthly_{chart_stock_name.replace(' ', '')}_{stock_code}_{current_date}.png"
    
    # 파일명에서 특수문자 제거 및 공백을 언더스코어로 변경
    base_filename = base_filename.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_")
    
    # 파일 중복 확인 및 버전 추가
    version = 1
    filename = base_filename
    filepath = os.path.join(charts_dir, filename)
    
    while os.path.exists(filepath):
        # 파일명에서 확장자 분리
        name_without_ext = base_filename.rsplit('.', 1)[0]
        ext = base_filename.rsplit('.', 1)[1]
        filename = f"{name_without_ext}_v{version}.{ext}"
        filepath = os.path.join(charts_dir, filename)
        version += 1
    
    # 차트 저장
    plt.savefig(filepath, dpi=100, bbox_inches='tight')
    print(f"💾 차트가 저장되었습니다: {filepath}")
    
    # 차트 뷰어를 띄우지 않고 차트 닫기
    plt.close(fig)  # 특정 figure 닫기
    plt.close('all')  # 모든 figure 닫기
    
    # 메모리 정리
    import gc
    gc.collect()
    
    # 차트 데이터 반환 (보조지표 포함) - 일봉 분석과 동일한 패턴
    return filepath, chart_stock_name, df

def get_stock_name(stock_code):
    """종목코드로 종목명을 가져오는 함수 - DB에서 조회 (일봉과 동일한 방식)"""
    try:
        # 데이터베이스 연결
        db = DatabaseManager()
        
        if not db.connect():
            print(f"   ⚠️ DB 연결 실패로 종목코드를 종목명으로 사용: {stock_code}")
            return stock_code
        
        # stocks 테이블에서 종목명 조회 (일봉과 동일한 간단한 방식)
        stock_name_query = "SELECT stock_name FROM stocks WHERE stock_code = %s"
        stock_info = db.fetch_one(stock_name_query, (stock_code,))
        
        if stock_info and stock_info['stock_name']:
            stock_name = stock_info['stock_name']
            print(f"   ✅ DB에서 종목명 조회 성공: {stock_code} -> {stock_name}")
            db.disconnect()
            return stock_name
        
        # 정확한 매칭이 실패한 경우, 대소문자 무시하고 조회
        stock_name_query_case_insensitive = "SELECT stock_name FROM stocks WHERE UPPER(stock_code) = UPPER(%s)"
        stock_info = db.fetch_one(stock_name_query_case_insensitive, (stock_code,))
        
        if stock_info and stock_info['stock_name']:
            stock_name = stock_info['stock_name']
            print(f"   ✅ DB에서 종목명 조회 성공 (대소문자 무시): {stock_code} -> {stock_name}")
            db.disconnect()
            return stock_name
        
        # 모든 시도가 실패한 경우
        print(f"   ⚠️ 종목코드 {stock_code}를 DB에서 찾을 수 없습니다.")
        print(f"   💡 stocks 테이블에 해당 종목이 등록되어 있는지 확인해주세요.")
        db.disconnect()
        return stock_code
            
    except Exception as e:
        print(f"   ⚠️ 종목명 조회 실패: {str(e)}")
        try:
            db.disconnect()
        except:
            pass
        return stock_code

def save_chart_data_to_json(chart_data, stock_code, stock_name, trading_type="거래량"):
    """차트 데이터를 JSON으로 저장 - Gemini AI 최적화"""
    if chart_data is None or chart_data.empty:
        print("❌ 저장할 차트 데이터가 없습니다.")
        return None
    
    try:
        print(f"\n📊 차트 데이터를 JSON으로 저장합니다...")
        
        # 시간대 정보 제거
        chart_data_clean = chart_data.copy()
        if chart_data_clean.index.tz is not None:
            chart_data_clean.index = chart_data_clean.index.tz_localize(None)
            print("   🔧 시간대 정보를 제거했습니다.")
        
        # JSON 저장 디렉토리 생성
        json_dir = "chart_data_json"
        if not os.path.exists(json_dir):
            os.makedirs(json_dir)
            print(f"📁 {json_dir} 폴더를 생성했습니다.")
        
        # 파일명 생성 (거래타입 포함)
        current_date = datetime.now().strftime("%Y%m%d")
        trading_type_short = "거래량" if trading_type == "거래량" else "거래대금"
        filename = f"monthly_{stock_name}_{stock_code}_{trading_type_short}_{current_date}.json"
        filename = filename.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_")
        filepath = os.path.join(json_dir, filename)
        
        # 중복 확인 (거래타입 포함된 파일명으로)
        version = 1
        while os.path.exists(filepath):
            name_without_ext = filename.rsplit('.', 1)[0]
            ext = filename.rsplit('.', 1)[1]
            filename = f"{name_without_ext}_v{version}.{ext}"
            filepath = os.path.join(json_dir, filename)
            version += 1
        
        # JSON 데이터 구조화
        json_data = {
            "metadata": {
                "stock_name": stock_name,
                "stock_code": stock_code,
                "created_at": datetime.now().isoformat(),
                "data_period": {
                    "start": chart_data_clean.index[0].strftime('%Y-%m-%d'),
                    "end": chart_data_clean.index[-1].strftime('%Y-%m-%d')
                },
                "total_records": len(chart_data_clean),
                "chart_type": "monthly"
            },
            "summary": {
                "latest_close": float(chart_data_clean['Close'].iloc[-1]),
                "latest_volume": int(chart_data_clean['Volume'].iloc[-1]),
                "price_change": float(chart_data_clean['Close'].iloc[-1] - chart_data_clean['Open'].iloc[0]),
                "price_change_pct": float(((chart_data_clean['Close'].iloc[-1] / chart_data_clean['Open'].iloc[0]) - 1) * 100),
                "highest_price": float(chart_data_clean['High'].max()),
                "lowest_price": float(chart_data_clean['Low'].min()),
                "avg_volume": float(chart_data_clean['Volume'].mean())
            },
            "technical_indicators": {
                "latest_values": {
                    # 이동평균선 (차트 표시용)
                    "ma6": float(chart_data_clean['MA6'].iloc[-1]) if 'MA6' in chart_data_clean else None,
                    "ma12": float(chart_data_clean['MA12'].iloc[-1]) if 'MA12' in chart_data_clean else None,
                    "ma24": float(chart_data_clean['MA24'].iloc[-1]) if 'MA24' in chart_data_clean else None,
                    # 보조지표 (차트 표시용)
                    "macd": float(chart_data_clean['MACD'].iloc[-1]) if 'MACD' in chart_data_clean else None,
                    "macd_signal": float(chart_data_clean['MACD_Signal'].iloc[-1]) if 'MACD_Signal' in chart_data_clean else None,
                    "macd_histogram": float(chart_data_clean['MACD_Histogram'].iloc[-1]) if 'MACD_Histogram' in chart_data_clean else None,
                    "rsi": float(chart_data_clean['RSI'].iloc[-1]) if 'RSI' in chart_data_clean else None,
                    # 볼린저밴드 (차트 표시용)
                    "bb_upper": float(chart_data_clean['BB_Upper'].iloc[-1]) if 'BB_Upper' in chart_data_clean else None,
                    "bb_middle": float(chart_data_clean['BB_Middle'].iloc[-1]) if 'BB_Middle' in chart_data_clean else None,
                    "bb_lower": float(chart_data_clean['BB_Lower'].iloc[-1]) if 'BB_Lower' in chart_data_clean else None
                }
            },
            "chart_data": []
        }
        
        # 차트 데이터 추가 (최근 30개 데이터만 - AI 분석에 충분)
        recent_data = chart_data_clean.tail(30)
        for date, row in recent_data.iterrows():
            data_point = {
                "date": date.strftime('%Y-%m-%d'),
                "open": float(row['Open']),
                "high": float(row['High']),
                "low": float(row['Low']),
                "close": float(row['Close']),
                "volume": int(row['Volume'])
            }
            
            # 기술적 지표 추가 (차트 표시용만)
            # 이동평균선
            if 'MA6' in row:
                data_point["ma6"] = float(row['MA6'])
            if 'MA12' in row:
                data_point["ma12"] = float(row['MA12'])
            if 'MA24' in row:
                data_point["ma24"] = float(row['MA24'])
            # 보조지표
            if 'MACD' in row:
                data_point["macd"] = float(row['MACD'])
            if 'MACD_Signal' in row:
                data_point["macd_signal"] = float(row['MACD_Signal'])
            if 'MACD_Histogram' in row:
                data_point["macd_histogram"] = float(row['MACD_Histogram'])
            if 'RSI' in row:
                data_point["rsi"] = float(row['RSI'])
            # 볼린저밴드
            if 'BB_Upper' in row:
                data_point["bb_upper"] = float(row['BB_Upper'])
            if 'BB_Middle' in row:
                data_point["bb_middle"] = float(row['BB_Middle'])
            if 'BB_Lower' in row:
                data_point["bb_lower"] = float(row['BB_Lower'])
            
            json_data["chart_data"].append(data_point)
        
        # JSON 파일 저장
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        print(f"💾 JSON 파일이 저장되었습니다: {filepath}")
        print(f"📊 데이터 구조:")
        print(f"   - 메타데이터: 종목 정보, 생성일시, 데이터 기간")
        print(f"   - 요약 정보: 최근 가격, 변동률, 거래량 통계")
        print(f"   - 기술적 지표: 최신 보조지표 값들")
        print(f"   - 차트 데이터: 최근 30개 거래월 OHLCV + 지표")
        
        return filepath
        
    except Exception as e:
        print(f"❌ JSON 파일 저장 중 오류: {e}")
        return None

def save_chart_data_to_csv(chart_data, stock_code, stock_name):
    """차트 데이터를 CSV로 저장 - 간단하고 읽기 쉬움"""
    if chart_data is None or chart_data.empty:
        print("❌ 저장할 차트 데이터가 없습니다.")
        return None
    
    try:
        print(f"\n📊 차트 데이터를 CSV로 저장합니다...")
        
        # 시간대 정보 제거
        chart_data_clean = chart_data.copy()
        if chart_data_clean.index.tz is not None:
            chart_data_clean.index = chart_data_clean.index.tz_localize(None)
            print("   🔧 시간대 정보를 제거했습니다.")
        
        # CSV 저장 디렉토리 생성
        csv_dir = "chart_data_csv"
        if not os.path.exists(csv_dir):
            os.makedirs(csv_dir)
            print(f"📁 {csv_dir} 폴더를 생성했습니다.")
        
        # 파일명 생성
        current_date = datetime.now().strftime("%Y%m%d")
        filename = f"monthly_{stock_name}_{stock_code}_{current_date}.csv"
        filename = filename.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_")
        filepath = os.path.join(csv_dir, filename)
        
        # 중복 확인
        version = 1
        while os.path.exists(filepath):
            name_without_ext = filename.rsplit('.', 1)[0]
            ext = filename.rsplit('.', 1)[1]
            filename = f"{name_without_ext}_v{version}.{ext}"
            filepath = os.path.join(csv_dir, filename)
            version += 1
        
        # CSV 저장 (최근 50개 데이터만)
        recent_data = chart_data_clean.tail(50)
        recent_data.to_csv(filepath, encoding='utf-8-sig')
        
        print(f"💾 CSV 파일이 저장되었습니다: {filepath}")
        print(f"📊 데이터: 최근 50개 거래월 OHLCV + 기술적 지표")
        
        return filepath
        
    except Exception as e:
        print(f"❌ CSV 파일 저장 중 오류: {e}")
        return None

def save_chart_summary_to_text(chart_data, stock_code, stock_name):
    """차트 데이터 요약을 텍스트로 저장 - AI 분석 최적화"""
    if chart_data is None or chart_data.empty:
        print("❌ 저장할 차트 데이터가 없습니다.")
        return None
    
    try:
        print(f"\n📊 차트 데이터 요약을 텍스트로 저장합니다...")
        
        # 텍스트 저장 디렉토리 생성
        text_dir = "chart_data_text"
        if not os.path.exists(text_dir):
            os.makedirs(text_dir)
            print(f"📁 {text_dir} 폴더를 생성했습니다.")
        
        # 파일명 생성
        current_date = datetime.now().strftime("%Y%m%d")
        filename = f"monthly_{stock_name}_{stock_code}_{current_date}_summary.txt"
        filename = filename.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_")
        filepath = os.path.join(text_dir, filename)
        
        # 중복 확인
        version = 1
        while os.path.exists(filepath):
            name_without_ext = filename.rsplit('.', 1)[0]
            ext = filename.rsplit('.', 1)[1]
            filename = f"{name_without_ext}_v{version}.{ext}"
            filepath = os.path.join(text_dir, filename)
            version += 1
        
        # 요약 텍스트 생성
        summary_text = f"""주식 월봉 차트 데이터 요약
========================

종목 정보:
- 종목명: {stock_name}
- 종목코드: {stock_code}
- 생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- 데이터 기간: {chart_data.index[0].strftime('%Y-%m-%d')} ~ {chart_data.index[-1].strftime('%Y-%m-%d')}
- 총 데이터 수: {len(chart_data)}개월

가격 정보:
- 시작가: {chart_data['Open'].iloc[0]:,.0f}원
- 최근 종가: {chart_data['Close'].iloc[-1]:,.0f}원
- 최고가: {chart_data['High'].max():,.0f}원
- 최저가: {chart_data['Low'].min():,.0f}원
- 가격 변동: {chart_data['Close'].iloc[-1] - chart_data['Open'].iloc[0]:+,.0f}원
- 변동률: {((chart_data['Close'].iloc[-1] / chart_data['Open'].iloc[0]) - 1) * 100:+.2f}%

거래량 정보:
- 평균 거래량: {chart_data['Volume'].mean():,.0f}주
- 최대 거래량: {chart_data['Volume'].max():,.0f}주
- 최근 거래량: {chart_data['Volume'].iloc[-1]:,.0f}주

기술적 지표 (최근값):
"""
        
        # 기술적 지표 추가
        if 'MA5' in chart_data:
            summary_text += f"- 5개월 이동평균: {chart_data['MA5'].iloc[-1]:,.0f}원\n"
        if 'MA10' in chart_data:
            summary_text += f"- 10개월 이동평균: {chart_data['MA10'].iloc[-1]:,.0f}원\n"
        if 'MA20' in chart_data:
            summary_text += f"- 20개월 이동평균: {chart_data['MA20'].iloc[-1]:,.0f}원\n"
        if 'MA60' in chart_data:
            summary_text += f"- 60개월 이동평균: {chart_data['MA60'].iloc[-1]:,.0f}원\n"
        if 'CCI' in chart_data:
            summary_text += f"- CCI: {chart_data['CCI'].iloc[-1]:.1f}\n"
        if 'ADX' in chart_data:
            summary_text += f"- ADX: {chart_data['ADX'].iloc[-1]:.1f}\n"
        if 'Plus_DI' in chart_data:
            summary_text += f"- +DI: {chart_data['Plus_DI'].iloc[-1]:.1f}\n"
        if 'Minus_DI' in chart_data:
            summary_text += f"- -DI: {chart_data['Minus_DI'].iloc[-1]:.1f}\n"
        
        summary_text += f"""
최근 10개 거래월 데이터:
"""
        
        # 최근 10개 데이터 추가
        recent_data = chart_data.tail(10)
        for date, row in recent_data.iterrows():
            summary_text += f"{date.strftime('%Y-%m-%d')}: {row['Open']:,.0f} → {row['Close']:,.0f} (거래량: {row['Volume']:,.0f})\n"
        
        # 텍스트 파일 저장
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(summary_text)
        
        print(f"💾 텍스트 요약 파일이 저장되었습니다: {filepath}")
        
        return filepath
        
    except Exception as e:
        print(f"❌ 텍스트 파일 저장 중 오류: {e}")
        return None

# 엑셀 저장 기능 주석 처리 (나중에 검토용으로 사용)
'''
def save_chart_data_to_excel(chart_data, stock_code, stock_name):
    """차트 데이터를 엑셀로 저장 (보조지표 포함)"""
    if chart_data is None or chart_data.empty:
        print("❌ 저장할 차트 데이터가 없습니다.")
        return None
    
    try:
        print(f"\n📊 차트 데이터를 엑셀로 저장합니다...")
        
        # 시간대 정보 제거 (Excel 호환성을 위해)
        chart_data_clean = chart_data.copy()
        if chart_data_clean.index.tz is not None:
            chart_data_clean.index = chart_data_clean.index.tz_localize(None)
            print("   🔧 시간대 정보를 제거했습니다.")
        
        # 엑셀 파일 저장 디렉토리 생성
        excel_dir = "chart_data_excel"
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)
            print(f"📁 {excel_dir} 폴더를 생성했습니다.")
        
        # 파일명 생성
        current_date = datetime.now().strftime("%Y%m%d")
        base_filename = f"monthly_chart_data_{stock_name}_{stock_code}_{current_date}.xlsx"
        base_filename = base_filename.replace(" ", "_").replace("/", "_").replace("\\", "_").replace(":", "_")
        
        # 파일 중복 확인 및 버전 추가
        version = 1
        filename = base_filename
        filepath = os.path.join(excel_dir, filename)
        
        while os.path.exists(filepath):
            name_without_ext = base_filename.rsplit('.', 1)[0]
            ext = base_filename.rsplit('.', 1)[1]
            filename = f"{name_without_ext}_v{version}.{ext}"
            filepath = os.path.join(excel_dir, filename)
            version += 1
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        
        # 기본 시트 제거
        wb.remove(wb.active)
        
        # 1. 종합 데이터 시트 (모든 지표 포함)
        ws_summary = wb.create_sheet("종합데이터")
        
        # 모든 컬럼 선택
        summary_data = chart_data_clean.copy()
        summary_data.index.name = 'Date'
        summary_data.insert(0, 'Date', summary_data.index.strftime('%Y-%m-%d'))
        
        for r in dataframe_to_rows(summary_data, index=False, header=True):
            ws_summary.append(r)
        
        # 헤더 스타일링
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 컬럼 너비 조정
        for col in ws_summary.columns:
            ws_summary.column_dimensions[col[0].column_letter].width = 12
        
        # 2. 요약 정보 시트
        ws_info = wb.create_sheet("요약정보")
        
        # 기본 정보
        info_data = [
            ["종목명", stock_name],
            ["종목코드", stock_code],
            ["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["데이터 기간", f"{chart_data_clean.index[0].strftime('%Y-%m-%d')} ~ {chart_data_clean.index[-1].strftime('%Y-%m-%d')}"],
            ["총 데이터 수", len(chart_data_clean)],
            ["", ""],
            ["최근 데이터 요약", ""],
            ["최근 종가", f"{chart_data_clean['Close'].iloc[-1]:,.0f}원"],
            ["최근 CCI", f"{chart_data_clean['CCI'].iloc[-1]:.1f}"],
            ["최근 ADX", f"{chart_data_clean['ADX'].iloc[-1]:.1f}"],
            ["5개월 이동평균", f"{chart_data_clean['MA5'].iloc[-1]:,.0f}원"],
            ["10개월 이동평균", f"{chart_data_clean['MA10'].iloc[-1]:,.0f}원"],
            ["20개월 이동평균", f"{chart_data_clean['MA20'].iloc[-1]:,.0f}원"],
            ["60개월 이동평균", f"{chart_data_clean['MA60'].iloc[-1]:,.0f}원"],
        ]
        
        for row in info_data:
            ws_info.append(row)
        
        # 헤더 스타일링
        for row in ws_info.iter_rows(min_row=1, max_row=len(info_data)):
            for cell in row:
                if cell.value and cell.value in ["종목명", "종목코드", "생성일시", "데이터 기간", "총 데이터 수", "최근 데이터 요약"]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 컬럼 너비 조정
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 30
        
        # 파일 저장
        wb.save(filepath)
        print(f"💾 엑셀 파일이 저장되었습니다: {filepath}")
        
        # 시트 정보 출력
        print(f"📊 생성된 시트:")
        print(f"   - 종합데이터: 모든 지표 통합 (OHLCV + 기술적 지표)")
        print(f"   - 요약정보: 종목 및 데이터 요약")
        
        return filepath
        
    except Exception as e:
        print(f"❌ 엑셀 파일 저장 중 오류: {e}")
        return None
'''

def get_monthly_stock_data(stock_code):
    """국내 주식 월봉 데이터 조회 (10년/120개월+) - DB에서 조회"""
    print(f"🔍 {stock_code} 10년(120개월+) 월봉 시세 조회 중...")
    print("   📅 월봉 데이터는 거래일 기준으로 제공되며, 월말 기준으로 집계됩니다.")
    
    # DB에서 월봉 데이터 조회 시도
    db_monthly_data = get_monthly_stock_data_from_db(stock_code)
    if db_monthly_data is not None and not db_monthly_data.empty:
        print(f"   ✅ DB에서 월봉 데이터 조회 완료")
        return db_monthly_data
    
    # DB에서 실패한 경우 오류 메시지 출력
    print(f"   ⚠️ DB에서 월봉 데이터 조회 실패")
    
    # 네이버 금융 데이터 조회 (우선)
    print("   🔄 네이버 금융에서 실시간 데이터 확인 중...")
    try:
        # from naver_data_module import get_naver_stock_data, get_naver_historical_data
        
        # naver_result = get_naver_stock_data(stock_code)
        # if naver_result['success']:
        #     print(f"   ✅ 네이버 금융 실시간 데이터: {naver_result['stock_name']}")
        #     print(f"   📈 현재가: {naver_result['current_price']:,.0f}원")
        #     print(f"   📊 변동: {naver_result['change_direction']} {naver_result['change_amount']:+,}원")
        #     print(f"   ⏰ 조회시간: {naver_result['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
        print("   ⚠️ 네이버 금융 모듈이 비활성화되었습니다.")
    except ImportError:
        print("   ⚠️ 네이버 금융 모듈을 불러올 수 없습니다.")
    
    # 모든 소스에서 실패
    print("❌ 월봉 데이터 조회에 실패했습니다.")
    print("💡 가능한 원인:")
    print("   - 종목코드가 잘못되었습니다")
    print("   - 해당 종목이 상장폐지되었습니다")
    print("   - DB에 월봉 데이터가 수집되지 않았습니다")
    print("   - 데이터베이스 연결에 문제가 있습니다")
    return None
    # Yahoo Finance 관련 코드 제거됨 - DB 전용으로 변경
    # if yf_monthly_data is not None:
    #     # 최신 데이터 확인 (현재 날짜와 비교)
    #     latest_monthly_date = yf_monthly_data.index[-1]
    #     # 타임존 정보 제거
    #     if hasattr(latest_monthly_date, 'tz_localize'):
    #         latest_monthly_date = latest_monthly_date.tz_localize(None)
    #     elif hasattr(latest_monthly_date, 'replace'):
    #         latest_monthly_date = latest_monthly_date.replace(tzinfo=None)
    #     
    #     current_date = datetime.now()
    #     days_diff = (current_date - latest_monthly_date).days
    #     
    #     print(f"   📅 Yahoo Finance 월봉 최신 데이터: {latest_monthly_date.strftime('%Y-%m-%d')}")
    #     print(f"   📅 현재 날짜: {current_date.strftime('%Y-%m-%d')}")
    #     print(f"   📅 데이터 차이: {days_diff}일")
    #     
    #     # 7일 이상 차이나면 일봉 데이터로 최신 월봉 보완
    #     if days_diff > 7:
    #         print(f"   ⚠️ Yahoo Finance 월봉 데이터가 {days_diff}일 전 데이터입니다.")
    #         print(f"   🔄 Yahoo Finance 일봉 데이터로 최신 월봉을 보완합니다...")
    #         
    #         # Yahoo Finance에서 일봉 데이터 조회 (최근 90일)
    #         try:
    #             daily_hist = stock.history(period="90d", interval="1d")
    #             if not daily_hist.empty:
    #                 print(f"   ✅ Yahoo Finance 일봉: {daily_hist.index[0].strftime('%Y-%m-%d')} ~ {daily_hist.index[-1].strftime('%Y-%m-%d')}")
    #                 print(f"   📊 일봉 데이터 상세:")
    #                 for i, (date, row) in enumerate(daily_hist.tail(5).iterrows()):
    #                     print(f"      {date.strftime('%Y-%m-%d')}: {row['Open']:,.0f} → {row['Close']:,.0f}")
    #                 
    #                 # 일봉을 월봉으로 변환
    #                 enhanced_monthly_data = convert_daily_to_monthly(daily_hist, yf_monthly_data)
    #                 if enhanced_monthly_data is not None:
    #                     print(f"   ✅ 일봉 데이터로 월봉을 보완했습니다!")
    #                     print(f"   📅 최신 월봉 데이터: {enhanced_monthly_data.index[-1].strftime('%Y-%m-%d')}")
    #                     return enhanced_monthly_data
    #                 else:
    #                     print(f"   ⚠️ 일봉 데이터 변환에 실패하여 기존 월봉 데이터를 사용합니다.")
    #             else:
    #                 print(f"   ⚠️ Yahoo Finance 일봉 데이터를 가져올 수 없어 기존 월봉 데이터를 사용합니다.")
    #         except Exception as e:
    #             print(f"   ❌ Yahoo Finance 일봉 데이터 조회 실패: {str(e)[:50]}...")
    #     
    #     return yf_monthly_data
    
    # Yahoo Finance에서 월봉 데이터를 가져올 수 없는 경우
    # print("   ⚠️ Yahoo Finance에서 월봉 데이터를 가져올 수 없습니다.")
    # print("   🔄 Yahoo Finance 일봉 데이터로 월봉을 생성합니다...")
    # 
    # # Yahoo Finance에서 일봉 데이터로 월봉 생성 시도
    # for ticker in tickers_to_try:
    #     try:
    #         stock = yf.Ticker(ticker)
    #         # 10년 일봉 데이터 조회
    #         daily_hist = stock.history(period="10y", interval="1d")
    #         if not daily_hist.empty:
    #             print(f"   ✅ Yahoo Finance 일봉: {daily_hist.index[0].strftime('%Y-%m-%d')} ~ {daily_hist.index[-1].strftime('%Y-%m-%d')}")
    #             
    #             # 일봉을 월봉으로 변환
    #             monthly_from_daily = convert_daily_to_monthly(daily_hist, None)
    #             if monthly_from_daily is not None:
    #                 print(f"   ✅ 일봉 데이터로 월봉을 생성했습니다!")
    #                 return monthly_from_daily
    #             break
    #     except Exception as e:
    #         print(f"   ❌ {ticker} 일봉 시도 실패: {str(e)[:50]}...")
    #         continue
    # 
    # # 모든 소스에서 실패
    # print("❌ 월봉 데이터 조회에 실패했습니다.")
    # print("💡 가능한 원인:")
    # print("   - 종목코드가 잘못되었습니다")
    # print("   - 해당 종목이 상장폐지되었습니다")
    # print("   - Yahoo Finance에서 지원하지 않는 종목입니다")
    # return None

def main():
    """메인 함수"""
    print("🚀 국내 주식 월봉 시세 조회 프로그램 (10년/120개월+)")
    print("="*60)
    
    # 종목코드 입력
    while True:
        stock_code = input("📈 종목코드를 입력하세요 (예: 005930): ").strip()
        if len(stock_code) == 6 and (stock_code.isdigit() or stock_code.isalnum()):
            break
        else:
            print("❌ 올바른 종목코드를 입력해주세요 (6자리 숫자 또는 영문+숫자)")
    
    # 월봉 데이터 조회
    hist = get_monthly_stock_data(stock_code)
    
    if hist is not None:
        # 월봉 데이터 분석
        analyze_monthly_stock_data(hist, stock_code)
        
        # 월봉 차트 생성 (차트 데이터 반환) - 일봉 분석과 동일한 패턴
        chart_result = create_monthly_stock_chart(hist, stock_code)
        
        if chart_result and len(chart_result) == 3:
            chart_path, stock_name, chart_data = chart_result
            print(f"🏢 종목명: {stock_name}")
            
            # JSON 저장 (추천)
            json_path = save_chart_data_to_json(chart_data, stock_code, stock_name)
            
            # CSV 저장 (보조)
            csv_path = save_chart_data_to_csv(chart_data, stock_code, stock_name)
            
            # 텍스트 요약 저장 (보조)
            text_path = save_chart_summary_to_text(chart_data, stock_code, stock_name)
            
            if json_path:
                print(f"\n✅ 월봉 분석이 완료되었습니다!")
                print(f"📈 차트 이미지: {chart_path}")
                print(f"📊 JSON 데이터: {json_path}")
                if csv_path:
                    print(f"📋 CSV 데이터: {csv_path}")
                if text_path:
                    print(f"📝 텍스트 요약: {text_path}")
                print(f"\n💡 AI 분석을 원하시면 다음 명령어를 사용하세요:")
                print(f"   from ai_chart_analysis import analyze_stock_chart")
                print(f"   result = analyze_stock_chart('{stock_code}', '월봉')")
            else:
                print(f"\n✅ 월봉 분석이 완료되었습니다!")
                print(f"📈 차트 이미지: {chart_path}")
                print(f"❌ 데이터 파일 저장에 실패했습니다.")
        else:
            print(f"\n❌ 차트 생성에 실패했습니다.")
    else:
        print("\n❌ 월봉 데이터 조회에 실패했습니다.")

if __name__ == "__main__":
    main() 